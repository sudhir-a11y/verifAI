import logging
import base64
import csv
import io
import json
import re
from datetime import date, datetime, timedelta, timezone
from uuid import UUID
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import PlainTextResponse, Response
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.deps.auth import require_roles
from app.core.config import settings
from app.db.session import SessionLocal, get_db
from app.schemas.auth import UserRole
from app.schemas.qc_tools import (
    CompletedReportLatestHtmlResponse,
    CompletedReportQcStatusRequest,
    CompletedReportQcStatusResponse,
    CompletedReportUploadStatusRequest,
    CompletedReportUploadStatusResponse,
    ExcelImportResponse,
)
from app.services.auth_service import AuthenticatedUser
from app.services.ml_claim_model import (
    AUDITOR_QC_LABEL_TYPE,
    MODEL_KEY,
    ensure_model,
    upsert_feedback_label,
)
from app.services.sql_dump_parser import iter_table_rows_from_sql_dump_bytes
from app.services.access_control import doctor_matches_assignment

router = APIRouter(prefix="/user-tools", tags=["user-tools"])

logger = logging.getLogger(__name__)
_EMPTY_LIKE_TEXT_VALUES = {
    "na",
    "n/a",
    "not available",
    "none",
    "nil",
    "null",
    "-",
    ".",
}


def _normalize_optional_text(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    if raw.lower() in _EMPTY_LIKE_TEXT_VALUES:
        return ""
    return raw


def _normalize_report_title_html(value: Any) -> str:
    html = str(value or "")
    if not html:
        return ""
    return html.replace("HEALTH CLAIM INVESTIGATION REPORT", "HEALTH CLAIM ASSESSMENT SHEET")

def _parse_datetime_utc(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        dt = value
    else:
        raw = str(value or "").strip()
        if not raw:
            return None
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except Exception:
            return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _latest_model_trained_recently(db: Session, min_minutes: int) -> bool:
    if min_minutes <= 0:
        return False
    try:
        row = db.execute(
            text(
                """
                SELECT COALESCE(effective_from, created_at) AS trained_at
                FROM model_registry
                WHERE model_key = :model_key
                ORDER BY
                    CASE WHEN status = 'active' THEN 0 ELSE 1 END,
                    COALESCE(effective_from, created_at) DESC,
                    created_at DESC
                LIMIT 1
                """
            ),
            {"model_key": MODEL_KEY},
        ).mappings().first()
    except Exception:
        return False
    if row is None:
        return False
    trained_at = _parse_datetime_utc(row.get("trained_at"))
    if trained_at is None:
        return False
    cutoff = datetime.now(timezone.utc) - timedelta(minutes=max(0, min_minutes))
    return trained_at >= cutoff


def _run_ml_retrain_background(triggered_by: str, claim_id: str) -> None:
    if not settings.ml_auto_retrain_on_qc_yes:
        return
    db = SessionLocal()
    try:
        min_interval = int(settings.ml_auto_retrain_min_interval_minutes or 0)
        if _latest_model_trained_recently(db, min_interval):
            logger.info(
                "Skipping ML retrain (throttled). trigger=%s claim_id=%s min_interval=%s",
                triggered_by,
                claim_id,
                min_interval,
            )
            return
        model = ensure_model(db=db, force_retrain=True)
        version = str(model.get("version") or "") if isinstance(model, dict) else ""
        examples = int(model.get("num_examples") or 0) if isinstance(model, dict) else 0
        logger.info(
            "ML retrain completed. trigger=%s claim_id=%s version=%s examples=%s",
            triggered_by,
            claim_id,
            version,
            examples,
        )
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        logger.exception(
            "ML retrain failed. trigger=%s claim_id=%s error=%s",
            triggered_by,
            claim_id,
            exc,
        )
    finally:
        db.close()



def _is_valid_date(value: str | None) -> bool:
    if not value:
        return False
    try:
        date.fromisoformat(value)
        return True
    except Exception:
        return False


def _normalize_doctor_key(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _coerce_claim_status(raw: str | None) -> str:
    v = (raw or "").strip().lower()
    valid = {
        "ready_for_assignment",
        "waiting_for_documents",
        "in_review",
        "needs_qc",
        "completed",
        "withdrawn",
    }
    return v if v in valid else "waiting_for_documents"


TAGGING_SUBTAGGING_OPTIONS: dict[str, list[str]] = {
    "Genuine": [
        "Hospitalization verified and found to be genuine",
    ],
    "Fraudulent": [
        "Non cooperation of Hospital / patient during investigation",
        "Circumstantial evidence suggesting of possible fraud",
        "Infalted bills",
        "OPD to IPD conversion",
    ],
}


def _normalize_tagging(value: str | None) -> str:
    normalized = str(value or "").strip().lower()
    if normalized == "genuine":
        return "Genuine"
    if normalized in {"fraudulent", "fraudlent"}:
        return "Fraudulent"
    return ""


def _normalize_subtagging(tagging: str, value: str | None) -> str:
    options = TAGGING_SUBTAGGING_OPTIONS.get(tagging, [])
    raw = str(value or "").strip()
    if not raw:
        return ""
    for option in options:
        if raw.lower() == option.lower():
            return option
    return ""



def _normalize_qc_status(value: str | None) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return "no"
    compact = raw.replace("-", "_").replace(" ", "_")
    if compact in {"yes", "qc_yes", "qcyes", "qc_done", "done"}:
        return "yes"
    if compact in {"no", "qc_no", "qcno", "pending", "not_done"}:
        return "no"
    return ""


def _recommendation_to_feedback_label(raw: str | None) -> str | None:
    recommendation = str(raw or "").strip().lower()
    if recommendation in {"approve", "approved", "admissible", "payable"}:
        return "approve"
    if recommendation in {"reject", "rejected", "inadmissible"}:
        return "reject"
    if recommendation in {"need_more_evidence", "query", "manual_review"}:
        return "need_more_evidence"
    return None

def _system_report_sql(column_expr: str) -> str:
    col = f"LOWER(COALESCE({column_expr}, ''))"
    return (
        f"({col} LIKE 'system:%' OR {col} IN "
        "('system', 'system_ml', 'system-ai', 'ml-system', 'checklist_pipeline'))"
    )


def _report_source_from_created_by(created_by: str | None) -> str:
    raw = str(created_by or '').strip().lower()
    if raw.startswith('system:') or raw in {'system', 'system_ml', 'system-ai', 'ml-system', 'checklist_pipeline'}:
        return 'system'
    return 'doctor'


def _ensure_claim_report_uploads_table(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS claim_report_uploads (
                id BIGSERIAL PRIMARY KEY,
                claim_id UUID NOT NULL UNIQUE REFERENCES claims(id) ON DELETE CASCADE,
                report_export_status VARCHAR(30) NOT NULL DEFAULT 'pending',
                tagging VARCHAR(120),
                subtagging VARCHAR(120),
                opinion TEXT,
                qc_status VARCHAR(10) NOT NULL DEFAULT 'no',
                updated_by VARCHAR(100),
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_claim_report_uploads_claim_id ON claim_report_uploads(claim_id)"))


def _ensure_claim_legacy_data_table(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS claim_legacy_data (
                id BIGSERIAL PRIMARY KEY,
                claim_id UUID NOT NULL UNIQUE REFERENCES claims(id) ON DELETE CASCADE,
                legacy_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_claim_legacy_data_claim_id ON claim_legacy_data(claim_id)"))



def _ensure_claim_completed_at_column(db: Session) -> None:
    db.execute(text("ALTER TABLE claims ADD COLUMN IF NOT EXISTS completed_at TIMESTAMPTZ"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_claims_completed_at ON claims(completed_at)"))
    db.execute(
        text(
            """
            WITH first_completed AS (
                SELECT
                    we.claim_id,
                    MIN(we.occurred_at) AS first_completed_at
                FROM workflow_events we
                WHERE we.event_type = 'claim_status_updated'
                  AND COALESCE(we.event_payload->>'status', '') = 'completed'
                GROUP BY we.claim_id
            )
            UPDATE claims c
            SET completed_at = fc.first_completed_at
            FROM first_completed fc
            WHERE c.id = fc.claim_id
              AND c.status = 'completed'
              AND c.completed_at IS NULL
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE claims
            SET completed_at = updated_at
            WHERE status = 'completed'
              AND completed_at IS NULL
              AND updated_at IS NOT NULL
            """
        )
    )

def _normalize_doctor_token(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _split_doctor_filter(raw: str | None) -> list[str]:
    tokens: list[str] = []
    for part in str(raw or "").split(","):
        normalized = _normalize_doctor_token(part)
        if normalized:
            tokens.append(normalized)
    return tokens

def _normalize_import_key(value: Any) -> str:
    key = str(value or "").strip().lower()
    if not key:
        return ""
    key = key.replace("\ufeff", "")
    key = re.sub(r"[\s\-/]+", "_", key)
    key = re.sub(r"[^a-z0-9_]", "", key)
    key = re.sub(r"_+", "_", key).strip("_")
    return key


def _row_with_normalized_aliases(raw_map: dict[Any, Any]) -> dict[str, str]:
    norm: dict[str, str] = {}
    for key, value in (raw_map or {}).items():
        raw_key = str(key or "").strip().lower()
        if not raw_key:
            continue
        text_value = str(value or "").strip()
        norm[raw_key] = text_value
        normalized_key = _normalize_import_key(raw_key)
        if normalized_key and normalized_key not in norm:
            norm[normalized_key] = text_value
    return norm


def _parse_csv_rows(payload: bytes) -> list[dict[str, str]]:
    text_payload = payload.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text_payload))
    rows: list[dict[str, str]] = []
    for row in reader:
        rows.append(_row_with_normalized_aliases(row))
    return rows


def _parse_xlsx_rows(payload: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed for xlsx import") from exc

    wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(h or "").strip().lower() for h in headers_raw]
    rows: list[dict[str, str]] = []
    for row_vals in rows_iter:
        raw_row: dict[str, str] = {}
        for idx, val in enumerate(row_vals):
            key = headers[idx] if idx < len(headers) else f"col_{idx+1}"
            raw_row[key] = str(val).strip() if val is not None else ""
        rows.append(_row_with_normalized_aliases(raw_row))
    return rows


def _parse_sql_dump_rows(payload: bytes) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for raw_row in iter_table_rows_from_sql_dump_bytes(payload, "excel_case_uploads"):
        prepared_row: dict[str, str] = {}
        for key, value in (raw_row or {}).items():
            raw_key = str(key or "").strip().lower()
            if not raw_key:
                continue
            if value is None:
                prepared_row[raw_key] = ""
            elif isinstance(value, bool):
                prepared_row[raw_key] = "1" if value else "0"
            else:
                prepared_row[raw_key] = str(value).strip()
        norm = _row_with_normalized_aliases(prepared_row)
        if norm:
            rows.append(norm)
    return rows


def _extract_claim_fields(row: dict[str, str]) -> dict[str, Any] | None:
    external_claim_id = (
        row.get("external_claim_id")
        or row.get("claim_id")
        or row.get("claim no")
        or row.get("claim")
        or ""
    ).strip()
    if not external_claim_id:
        return None

    patient_name = (
        row.get("patient_name")
        or row.get("benef_name")
        or row.get("patient")
        or ""
    ).strip() or None

    patient_identifier = (
        row.get("patient_identifier")
        or row.get("policy_number")
        or row.get("policy no")
        or ""
    ).strip() or None

    assigned_doctor_id = (
        row.get("assigned_doctor_id")
        or row.get("doctor_username")
        or row.get("doctor")
        or ""
    ).strip() or None

    status = _coerce_claim_status(row.get("status"))
    if assigned_doctor_id and status == "waiting_for_documents":
        status = "in_review"

    tags: list[str] = []
    for key in ["claim_type", "policy_type", "diagnosis", "hospital_name", "treatment_type"]:
        val = (row.get(key) or "").strip()
        if val:
            tags.append(val)

    return {
        "external_claim_id": external_claim_id,
        "patient_name": patient_name,
        "patient_identifier": patient_identifier,
        "status": status,
        "assigned_doctor_id": assigned_doctor_id,
        "priority": 3,
        "source_channel": "excel_import",
        "tags": tags,
    }


@router.post("/upload-excel", response_model=ExcelImportResponse)
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
) -> ExcelImportResponse:
    _ensure_claim_legacy_data_table(db)
    _ensure_claim_report_uploads_table(db)
    _ensure_claim_completed_at_column(db)

    payload = await file.read()
    if not payload:
        raise HTTPException(status_code=400, detail="empty file")

    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        rows = _parse_xlsx_rows(payload)
    elif filename.endswith(".sql"):
        rows = _parse_sql_dump_rows(payload)
    else:
        rows = _parse_csv_rows(payload)

    inserted = 0
    updated = 0
    skipped = 0
    rejected_rows: list[dict[str, Any]] = []

    for idx, row in enumerate(rows, start=2):
        claim = _extract_claim_fields(row)
        if claim is None:
            skipped += 1
            raw_claim_id = (
                str(
                    row.get("external_claim_id")
                    or row.get("claim_id")
                    or row.get("claim no")
                    or row.get("claim")
                    or ""
                ).strip()
            )
            rejected_rows.append(
                {
                    "row_number": idx,
                    "claim_id": raw_claim_id,
                    "reason": "Missing claim ID (external_claim_id / claim_id / claim no / claim).",
                }
            )
            continue

        try:
            with db.begin_nested():
                existing = db.execute(
                    text("SELECT id FROM claims WHERE external_claim_id = :external_claim_id LIMIT 1"),
                    {"external_claim_id": claim["external_claim_id"]},
                ).mappings().first()

                claim_uuid: str
                if existing is None:
                    inserted_row = db.execute(
                        text(
                            """
                            INSERT INTO claims (
                                external_claim_id, patient_name, patient_identifier, status,
                                assigned_doctor_id, priority, source_channel, tags, completed_at
                            ) VALUES (
                                :external_claim_id, :patient_name, :patient_identifier, CAST(:status AS claim_status),
                                :assigned_doctor_id, :priority, :source_channel, CAST(:tags AS jsonb),
                                CASE WHEN CAST(:status AS claim_status) = 'completed'::claim_status THEN NOW() ELSE NULL END
                            )
                            RETURNING id
                            """
                        ),
                        {
                            **claim,
                            "tags": json.dumps(claim["tags"]),
                        },
                    ).mappings().one()
                    claim_uuid = str(inserted_row["id"])
                    inserted += 1
                else:
                    skipped += 1
                    rejected_rows.append(
                        {
                            "row_number": idx,
                            "claim_id": str(claim.get("external_claim_id") or "").strip(),
                            "reason": "Claim ID already exists. Duplicate claim numbers are rejected.",
                        }
                    )
                    continue

                legacy_payload = dict(row)
                if not str(legacy_payload.get("claim_id") or "").strip():
                    legacy_payload["claim_id"] = str(claim.get("external_claim_id") or "")
                if not str(legacy_payload.get("source_file_name") or "").strip():
                    legacy_payload["source_file_name"] = str(file.filename or "")
                if not str(legacy_payload.get("uploaded_by_username") or "").strip():
                    legacy_payload["uploaded_by_username"] = str(current_user.username or "")

                db.execute(
                    text(
                        """
                        INSERT INTO claim_legacy_data (claim_id, legacy_payload, updated_at)
                        VALUES (:claim_id, CAST(:legacy_payload AS jsonb), NOW())
                        ON CONFLICT (claim_id)
                        DO UPDATE SET
                            legacy_payload = EXCLUDED.legacy_payload,
                            updated_at = NOW()
                        """
                    ),
                    {
                        "claim_id": claim_uuid,
                        "legacy_payload": json.dumps(legacy_payload),
                    },
                )
        except Exception as exc:
            skipped += 1
            reason_text = "Upload row failed"
            if str(exc).strip():
                reason_text = str(exc).replace("\r", " ").replace("\n", " ").strip()
            if len(reason_text) > 260:
                reason_text = reason_text[:257] + "..."
            rejected_rows.append(
                {
                    "row_number": idx,
                    "claim_id": str(claim.get("external_claim_id") or "").strip(),
                    "reason": reason_text,
                }
            )

    rejected_excel_base64 = ""
    rejected_excel_filename = ""
    if rejected_rows:
        try:
            from openpyxl import Workbook
            from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

            def _xlsx_text(v: Any) -> str:
                txt = str(v or "")
                txt = ILLEGAL_CHARACTERS_RE.sub("", txt)
                if len(txt) > 32767:
                    txt = txt[:32767]
                return txt

            wb = Workbook()
            ws = wb.active
            ws.title = "Rejected Rows"
            ws.append(["row_number", "claim_id", "reason"])
            for item in rejected_rows:
                ws.append(
                    [
                        int(item.get("row_number") or 0),
                        _xlsx_text(item.get("claim_id") or ""),
                        _xlsx_text(item.get("reason") or ""),
                    ]
                )
            out = io.BytesIO()
            wb.save(out)
            rejected_excel_base64 = base64.b64encode(out.getvalue()).decode("ascii")
            safe_name = str(file.filename or "upload").strip() or "upload"
            safe_name = re.sub(r"[\\/:*?\"<>|]+", "_", safe_name)
            safe_name = re.sub(r"\.[A-Za-z0-9]+$", "", safe_name)
            rejected_excel_filename = f"{safe_name}_rejected_rows.xlsx"
        except Exception as exc:
            logger.warning("failed to build rejected rows xlsx: %s", exc)

    db.commit()
    uploaded = int(inserted + updated)
    return ExcelImportResponse(
        total_rows=len(rows),
        uploaded=uploaded,
        inserted=inserted,
        updated=updated,
        skipped=skipped,
        rejected_count=len(rejected_rows),
        rejected_rows=rejected_rows,
        rejected_excel_base64=rejected_excel_base64,
        rejected_excel_filename=rejected_excel_filename,
    )

@router.get("/completed-reports")
def completed_reports(
    status_filter: str = Query(default="pending"),
    qc_filter: str = Query(default="no"),
    search_claim: str | None = Query(default=None),
    allotment_date: str | None = Query(default=None),
    doctor_filter: str | None = Query(default=None),
    exclude_tagged: bool = Query(default=False),
    sort_order: str = Query(default="updated_desc"),
    limit: int = Query(default=200, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user, UserRole.auditor, UserRole.doctor)),
) -> dict:
    _ensure_claim_report_uploads_table(db)

    _ensure_claim_legacy_data_table(db)

    normalized_status = status_filter if status_filter in {"pending", "uploaded", "all"} else "pending"
    normalized_qc = "all" if str(qc_filter or "").strip().lower() == "all" else (_normalize_qc_status(qc_filter) or "no")
    normalized_sort = str(sort_order or "updated_desc").strip().lower()
    if normalized_sort not in {"updated_desc", "allotment_asc"}:
        normalized_sort = "updated_desc"

    system_report_expr_latest = _system_report_sql("rv.created_by")
    system_report_expr_stats = _system_report_sql("created_by")
    allotment_date_expr = (
        "COALESCE(la.allotment_date, "
        "CASE "
        "WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{4}-\d{2}-\d{2}$' "
        "THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'YYYY-MM-DD') "
        "WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{2}-\d{2}-\d{4}$' "
        "THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD-MM-YYYY') "
        "WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{2}/\d{2}/\d{4}$' "
        "THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD/MM/YYYY') "
        "ELSE NULL END)"
    )
    if normalized_sort == "allotment_asc":
        order_by_sql = f"CASE WHEN {allotment_date_expr} IS NULL THEN 1 ELSE 0 END ASC, {allotment_date_expr} ASC, c.updated_at ASC"
    else:
        order_by_sql = "c.updated_at DESC"


    filters: list[str] = ["c.status = 'completed'"]
    params: dict[str, Any] = {"limit": limit, "offset": offset}

    if search_claim and search_claim.strip():
        filters.append("LOWER(c.external_claim_id) LIKE :search_claim")
        params["search_claim"] = f"%{search_claim.strip().lower()}%"

    if _is_valid_date(allotment_date):
        filters.append(f"{allotment_date_expr} = :allotment_date")
        params["allotment_date"] = allotment_date

    doctors = _split_doctor_filter(doctor_filter)
    if doctors:
        doctor_clauses: list[str] = []
        for idx, doctor in enumerate(doctors):
            key = f"doctor_{idx}"
            params[key] = doctor
            doctor_clauses.append(
                f":{key} = ANY(string_to_array(regexp_replace(LOWER(COALESCE(c.assigned_doctor_id, '')), '[^a-z0-9,]+', '', 'g'), ','))"
            )
        filters.append("(" + " OR ".join(doctor_clauses) + ")")

    effective_status_expr = """CASE
        WHEN NULLIF(TRIM(COALESCE(um.tagging, '')), '') IS NOT NULL
             OR NULLIF(TRIM(COALESCE(um.subtagging, '')), '') IS NOT NULL
             OR NULLIF(TRIM(COALESCE(um.opinion, '')), '') IS NOT NULL
        THEN 'uploaded'
        WHEN COALESCE(um.report_export_status, 'pending') = 'uploaded'
        THEN 'uploaded'
        WHEN COALESCE(rv.export_uri, '') <> ''
        THEN 'uploaded'
        ELSE 'pending'
    END"""

    where_sql = "WHERE " + " AND ".join(filters)
    status_where = ""
    if normalized_status != "all":
        status_where = f" AND {effective_status_expr} = :status_filter"
        params["status_filter"] = normalized_status

    qc_where = ""
    qc_expr = "CASE WHEN LOWER(REPLACE(REPLACE(COALESCE(um.qc_status, 'no'), ' ', '_'), '-', '_')) IN ('yes', 'qc_yes', 'qcyes', 'qc_done', 'done') THEN 'yes' ELSE 'no' END"
    if normalized_qc != "all":
        qc_where = f" AND {qc_expr} = :qc_filter"
        params["qc_filter"] = normalized_qc

    tagged_where = ""
    if exclude_tagged:
        tagged_where = " AND NULLIF(TRIM(COALESCE(um.tagging, '')), '') IS NULL"

    total = db.execute(
        text(
            f"""
            WITH latest_report AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id, report_status, report_markdown, export_uri, version_no, created_at, created_by
                FROM report_versions
                ORDER BY claim_id, version_no DESC
            ),
            latest_assignment AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    DATE(occurred_at) AS allotment_date
                FROM workflow_events
                WHERE event_type = 'claim_assigned'
                ORDER BY claim_id, occurred_at DESC
            ),
            upload_meta AS (
                SELECT
                    claim_id,
                    report_export_status,
                    tagging,
                    subtagging,
                    opinion,
                    qc_status,
                    updated_at
                FROM claim_report_uploads
            ),
            legacy_data AS (
                SELECT claim_id, legacy_payload
                FROM claim_legacy_data
            )
            SELECT COUNT(*)
            FROM claims c
            LEFT JOIN latest_report rv ON rv.claim_id = c.id
            LEFT JOIN latest_assignment la ON la.claim_id = c.id
            LEFT JOIN upload_meta um ON um.claim_id = c.id
            LEFT JOIN legacy_data ldata ON ldata.claim_id = c.id
            {where_sql}
            {status_where}
            {qc_where}
            {tagged_where}
            """
        ),
        params,
    ).scalar_one()

    rows = db.execute(
        text(
            f"""
            WITH latest_report AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id, report_status, report_markdown, export_uri, version_no, created_at, created_by
                FROM report_versions
                ORDER BY claim_id, version_no DESC
            ),
            latest_assignment AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    DATE(occurred_at) AS allotment_date
                FROM workflow_events
                WHERE event_type = 'claim_assigned'
                ORDER BY claim_id, occurred_at DESC
            ),
            report_counts AS (
                SELECT claim_id, COUNT(*) AS report_count
                FROM report_versions
                GROUP BY claim_id
            ),
            report_source_stats AS (
                SELECT
                    claim_id,
                    MAX(CASE WHEN {system_report_expr_stats} THEN 1 ELSE 0 END) AS has_system_html,
                    MAX(CASE WHEN NOT ({system_report_expr_stats}) THEN 1 ELSE 0 END) AS has_doctor_html
                FROM report_versions
                WHERE NULLIF(TRIM(COALESCE(report_markdown, '')), '') IS NOT NULL
                GROUP BY claim_id
            ),
            upload_meta AS (
                SELECT
                    claim_id,
                    report_export_status,
                    tagging,
                    subtagging,
                    opinion,
                    qc_status,
                    updated_at
                FROM claim_report_uploads
            ),
            legacy_data AS (
                SELECT claim_id, legacy_payload
                FROM claim_legacy_data
            )
            SELECT
                c.id,
                c.external_claim_id,
                c.patient_name,
                c.assigned_doctor_id,
                c.updated_at,
                c.updated_at AS completed_at,
                {allotment_date_expr} AS allotment_date,
                COALESCE(rv.report_status, 'pending') AS report_status,
                COALESCE(rv.export_uri, '') AS export_uri,
                rv.created_at AS report_created_at,
                COALESCE(rv.version_no, 0) AS report_version,
                CASE WHEN NULLIF(TRIM(COALESCE(rv.report_markdown, '')), '') IS NULL THEN FALSE ELSE TRUE END AS report_html_available,
                CASE WHEN {system_report_expr_latest} THEN 'system' ELSE 'doctor' END AS latest_report_source,
                CASE WHEN COALESCE(rss.has_doctor_html, 0) = 1 THEN TRUE ELSE FALSE END AS doctor_report_html_available,
                CASE WHEN COALESCE(rss.has_system_html, 0) = 1 THEN TRUE ELSE FALSE END AS system_report_html_available,
                COALESCE(um.report_export_status, 'pending') AS stored_report_export_status,
                COALESCE(um.tagging, '') AS tagging,
                COALESCE(um.subtagging, '') AS subtagging,
                COALESCE(um.opinion, '') AS opinion,
                CASE WHEN LOWER(REPLACE(REPLACE(COALESCE(um.qc_status, 'no'), ' ', '_'), '-', '_')) IN ('yes', 'qc_yes', 'qcyes', 'qc_done', 'done') THEN 'yes' ELSE 'no' END AS qc_status,
                um.updated_at AS upload_updated_at,
                COALESCE(rc.report_count, 0) AS report_count,
                {effective_status_expr} AS effective_report_status
            FROM claims c
            LEFT JOIN latest_report rv ON rv.claim_id = c.id
            LEFT JOIN latest_assignment la ON la.claim_id = c.id
            LEFT JOIN report_counts rc ON rc.claim_id = c.id
            LEFT JOIN report_source_stats rss ON rss.claim_id = c.id
            LEFT JOIN upload_meta um ON um.claim_id = c.id
            LEFT JOIN legacy_data ldata ON ldata.claim_id = c.id
            {where_sql}
            {status_where}
            {qc_where}
            {tagged_where}
            ORDER BY {order_by_sql}
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).mappings().all()

    items = []
    for row in rows:
        export_uri = str(row.get("export_uri") or "")
        effective_status = str(row.get("effective_report_status") or "pending")
        items.append(
            {
                "claim_uuid": str(row["id"]),
                "external_claim_id": str(row["external_claim_id"]),
                "patient_name": str(row.get("patient_name") or ""),
                "assigned_doctor_id": str(row.get("assigned_doctor_id") or ""),
                "report_status": effective_status,
                "effective_report_status": effective_status,
                "report_uploaded": effective_status == "uploaded" or bool(export_uri),
                "export_uri": export_uri,
                "updated_at": row.get("updated_at"),
                "completed_at": row.get("completed_at"),
                "allotment_date": row.get("allotment_date"),
                "report_created_at": row.get("report_created_at"),
                "report_version": int(row.get("report_version") or 0),
                "report_html_available": bool(row.get("report_html_available") or False),
                "latest_report_source": str(row.get("latest_report_source") or "doctor"),
                "doctor_report_html_available": bool(row.get("doctor_report_html_available") or False),
                "system_report_html_available": bool(row.get("system_report_html_available") or False),
                "report_count": int(row.get("report_count") or 0),
                "tagging": _normalize_optional_text(row.get("tagging")),
                "subtagging": _normalize_optional_text(row.get("subtagging")),
                "opinion": _normalize_optional_text(row.get("opinion")),
                "qc_status": str(row.get("qc_status") or "no"),
                "upload_updated_at": row.get("upload_updated_at"),
            }
        )

    return {
        "total": int(total or 0),
        "items": items,
        "tagging_subtagging_options": TAGGING_SUBTAGGING_OPTIONS,
    }


@router.post("/completed-reports/{claim_id}/upload-status", response_model=CompletedReportUploadStatusResponse)
def update_completed_report_upload_status(
    claim_id: UUID,
    payload: CompletedReportUploadStatusRequest,
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
) -> CompletedReportUploadStatusResponse:
    _ensure_claim_report_uploads_table(db)

    target_status = (payload.report_export_status or "uploaded").strip().lower()
    tagging = _normalize_tagging(payload.tagging)
    subtagging = _normalize_subtagging(tagging, payload.subtagging)
    opinion = str(payload.opinion or "").strip()

    if target_status != "uploaded":
        raise HTTPException(status_code=400, detail="Please select Uploaded status before saving.")
    if not tagging or not subtagging or not opinion:
        raise HTTPException(status_code=400, detail="Tagging, Subtagging and Opinion are mandatory.")

    claim = db.execute(
        text(
            """
            SELECT id, external_claim_id
            FROM claims
            WHERE id = :claim_id
              AND status = 'completed'
            LIMIT 1
            """
        ),
        {"claim_id": str(claim_id)},
    ).mappings().first()

    if claim is None:
        raise HTTPException(status_code=404, detail="Completed claim not found.")

    row = db.execute(
        text(
            """
            INSERT INTO claim_report_uploads (
                claim_id,
                report_export_status,
                tagging,
                subtagging,
                opinion,
                updated_by,
                updated_at
            )
            VALUES (
                :claim_id,
                :report_export_status,
                :tagging,
                :subtagging,
                :opinion,
                :updated_by,
                NOW()
            )
            ON CONFLICT (claim_id)
            DO UPDATE SET
                report_export_status = EXCLUDED.report_export_status,
                tagging = EXCLUDED.tagging,
                subtagging = EXCLUDED.subtagging,
                opinion = EXCLUDED.opinion,
                updated_by = EXCLUDED.updated_by,
                updated_at = NOW()
            RETURNING claim_id, report_export_status, tagging, subtagging, opinion, updated_at
            """
        ),
        {
            "claim_id": str(claim_id),
            "report_export_status": "uploaded",
            "tagging": tagging,
            "subtagging": subtagging,
            "opinion": opinion,
            "updated_by": current_user.username,
        },
    ).mappings().one()

    db.execute(
        text(
            """
            INSERT INTO workflow_events (claim_id, actor_type, actor_id, event_type, event_payload)
            VALUES (:claim_id, 'user', :actor_id, 'completed_report_upload_status_updated', CAST(:event_payload AS jsonb))
            """
        ),
        {
            "claim_id": str(claim_id),
            "actor_id": current_user.username,
            "event_payload": json.dumps(
                {
                    "report_export_status": "uploaded",
                    "tagging": tagging,
                    "subtagging": subtagging,
                }
            ),
        },
    )

    db.commit()

    return CompletedReportUploadStatusResponse(
        claim_id=str(row.get("claim_id") or claim_id),
        external_claim_id=str(claim.get("external_claim_id") or ""),
        report_export_status=str(row.get("report_export_status") or "uploaded"),
        tagging=_normalize_optional_text(row.get("tagging")),
        subtagging=_normalize_optional_text(row.get("subtagging")),
        opinion=_normalize_optional_text(row.get("opinion")),
        updated_at=str(row.get("updated_at") or ""),
    )


@router.post("/completed-reports/{claim_id}/qc-status", response_model=CompletedReportQcStatusResponse)
def update_completed_report_qc_status(
    claim_id: UUID,
    payload: CompletedReportQcStatusRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.auditor)),
) -> CompletedReportQcStatusResponse:
    _ensure_claim_report_uploads_table(db)

    qc_status = _normalize_qc_status(payload.qc_status) or ""
    if qc_status not in {"yes", "no"}:
        raise HTTPException(status_code=400, detail="Invalid QC status selected.")

    claim = db.execute(
        text(
            """
            SELECT id, external_claim_id
            FROM claims
            WHERE id = :claim_id
              AND status = 'completed'
            LIMIT 1
            """
        ),
        {"claim_id": str(claim_id)},
    ).mappings().first()

    if claim is None:
        raise HTTPException(status_code=404, detail="Completed claim not found.")

    row = db.execute(
        text(
            """
            INSERT INTO claim_report_uploads (
                claim_id,
                qc_status,
                updated_by,
                updated_at
            )
            VALUES (
                :claim_id,
                :qc_status,
                :updated_by,
                NOW()
            )
            ON CONFLICT (claim_id)
            DO UPDATE SET
                qc_status = EXCLUDED.qc_status,
                updated_by = EXCLUDED.updated_by,
                updated_at = NOW()
            RETURNING claim_id, qc_status, updated_at
            """
        ),
        {
            "claim_id": str(claim_id),
            "qc_status": qc_status,
            "updated_by": current_user.username,
        },
    ).mappings().one()

    feedback_label_value = None
    feedback_decision_id = None
    latest_decision = db.execute(
        text(
            """
            SELECT id, recommendation
            FROM decision_results
            WHERE claim_id = :claim_id
            ORDER BY generated_at DESC
            LIMIT 1
            """
        ),
        {"claim_id": str(claim_id)},
    ).mappings().first()

    if latest_decision is not None:
        feedback_decision_id = str(latest_decision.get("id") or "") or None
        if qc_status == "yes":
            feedback_label_value = _recommendation_to_feedback_label(str(latest_decision.get("recommendation") or ""))
            feedback_reason = "qc_status_marked_yes"
            feedback_notes = "Auto label captured when auditor marked QC as yes."
        else:
            feedback_label_value = "manual_review"
            feedback_reason = "qc_status_marked_no"
            feedback_notes = "Auto label captured when auditor marked QC as no."

        if feedback_label_value:
            try:
                upsert_feedback_label(
                    db=db,
                    claim_id=str(claim_id),
                    decision_id=feedback_decision_id,
                    label_type=AUDITOR_QC_LABEL_TYPE,
                    label_value=feedback_label_value,
                    override_reason=feedback_reason,
                    notes=feedback_notes,
                    created_by=current_user.username,
                )
            except Exception:
                logger.exception(
                    "Failed to upsert auditor QC feedback label. claim_id=%s qc_status=%s actor=%s",
                    str(claim_id),
                    qc_status,
                    current_user.username,
                )
                feedback_label_value = None

    event_payload = {"qc_status": qc_status}
    if feedback_label_value:
        event_payload["feedback_label"] = feedback_label_value
        event_payload["feedback_decision_id"] = feedback_decision_id

    db.execute(
        text(
            """
            INSERT INTO workflow_events (claim_id, actor_type, actor_id, event_type, event_payload)
            VALUES (:claim_id, 'user', :actor_id, 'completed_report_qc_status_updated', CAST(:event_payload AS jsonb))
            """
        ),
        {
            "claim_id": str(claim_id),
            "actor_id": current_user.username,
            "event_payload": json.dumps(event_payload),
        },
    )
    db.commit()

    if settings.ml_auto_retrain_on_qc_yes:
        background_tasks.add_task(_run_ml_retrain_background, current_user.username, str(claim_id))
        logger.info(
            "Queued ML retrain after QC audit update. claim_id=%s actor=%s qc_status=%s",
            str(claim_id),
            current_user.username,
            qc_status,
        )

    return CompletedReportQcStatusResponse(
        claim_id=str(row.get("claim_id") or claim_id),
        external_claim_id=str(claim.get("external_claim_id") or ""),
        qc_status=str(row.get("qc_status") or qc_status),
        updated_at=str(row.get("updated_at") or ""),
    )


@router.get("/completed-reports/{claim_id}/latest-html", response_model=CompletedReportLatestHtmlResponse)
def get_completed_report_latest_html(
    claim_id: UUID,
    source: str = Query(default="any"),
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user, UserRole.auditor, UserRole.doctor)),
) -> CompletedReportLatestHtmlResponse:
    normalized_source = str(source or "any").strip().lower() or "any"
    if normalized_source not in {"any", "doctor", "system"}:
        raise HTTPException(status_code=400, detail="invalid source. allowed: any, doctor, system")

    claim_meta = db.execute(
        text("SELECT assigned_doctor_id FROM claims WHERE id = :claim_id"),
        {"claim_id": str(claim_id)},
    ).mappings().first()
    if claim_meta is None:
        raise HTTPException(status_code=404, detail="claim not found")

    if current_user.role == UserRole.doctor and not doctor_matches_assignment(
        str(claim_meta.get("assigned_doctor_id") or ""),
        current_user.username,
    ):
        raise HTTPException(status_code=403, detail="doctor can access only assigned claims")

    system_report_expr = _system_report_sql("rv.created_by")
    source_where = ""
    if normalized_source == "doctor":
        source_where = f" AND NOT ({system_report_expr})"
    elif normalized_source == "system":
        source_where = f" AND {system_report_expr}"

    row = db.execute(
        text(
            f"""
            SELECT
                c.id AS claim_id,
                c.external_claim_id,
                rv.version_no,
                COALESCE(rv.report_markdown, '') AS report_html,
                COALESCE(rv.report_status, 'draft') AS report_status,
                COALESCE(rv.created_by, '') AS created_by,
                CASE WHEN {system_report_expr} THEN 'system' ELSE 'doctor' END AS report_source,
                rv.created_at
            FROM claims c
            JOIN report_versions rv ON rv.claim_id = c.id
            WHERE c.id = :claim_id
              AND NULLIF(TRIM(COALESCE(rv.report_markdown, '')), '') IS NOT NULL
            {source_where}
            ORDER BY rv.version_no DESC
            LIMIT 1
            """
        ),
        {"claim_id": str(claim_id)},
    ).mappings().first()

    report_html = _normalize_report_title_html(row.get("report_html")) if row is not None else ""
    if row is None or not report_html.strip():
        decision_system_report_expr = _system_report_sql("dr.generated_by")
        decision_source_where = ""
        if normalized_source == "doctor":
            decision_source_where = f" AND NOT ({decision_system_report_expr})"
        elif normalized_source == "system":
            decision_source_where = f" AND {decision_system_report_expr}"

        row = db.execute(
            text(
                f"""
                SELECT
                    c.id AS claim_id,
                    c.external_claim_id,
                    0 AS version_no,
                    NULLIF(TRIM(COALESCE(dr.decision_payload ->> 'report_html', '')), '') AS report_html,
                    COALESCE(NULLIF(TRIM(COALESCE(dr.decision_payload ->> 'report_status', '')), ''), 'draft') AS report_status,
                    COALESCE(dr.generated_by, '') AS created_by,
                    CASE WHEN {decision_system_report_expr} THEN 'system' ELSE 'doctor' END AS report_source,
                    dr.generated_at AS created_at
                FROM claims c
                JOIN decision_results dr ON dr.claim_id = c.id
                WHERE c.id = :claim_id
                  AND NULLIF(TRIM(COALESCE(dr.decision_payload ->> 'report_html', '')), '') IS NOT NULL
                  {decision_source_where}
                ORDER BY dr.generated_at DESC
                LIMIT 1
                """
            ),
            {"claim_id": str(claim_id)},
        ).mappings().first()

    if row is None:
        detail = (
            "No saved report HTML found for this claim and source."
            if normalized_source != "any"
            else "No saved report HTML found for this claim."
        )
        raise HTTPException(status_code=404, detail=detail)

    report_html = _normalize_report_title_html(row.get("report_html"))
    if not report_html.strip():
        detail = (
            "No saved report HTML found for this claim and source."
            if normalized_source != "any"
            else "No saved report HTML found for this claim."
        )
        raise HTTPException(status_code=404, detail=detail)

    return CompletedReportLatestHtmlResponse(
        claim_id=str(row.get("claim_id") or claim_id),
        external_claim_id=str(row.get("external_claim_id") or ""),
        version_no=int(row.get("version_no") or 0),
        report_html=report_html,
        report_status=str(row.get("report_status") or "draft"),
        report_source=str(row.get("report_source") or _report_source_from_created_by(row.get("created_by"))),
        created_by=str(row.get("created_by") or ""),
        created_at=str(row.get("created_at") or ""),
    )

@router.get("/allotment-date-wise")
def allotment_date_wise(
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
    db: Session = Depends(get_db),
    _current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
) -> dict:
    _ensure_claim_legacy_data_table(db)
    _ensure_claim_report_uploads_table(db)
    system_report_expr = _system_report_sql("created_by")

    filters = []
    params: dict[str, Any] = {}

    if _is_valid_date(from_date):
        filters.append("b.allotment_date >= :from_date")
        params["from_date"] = from_date
    if _is_valid_date(to_date):
        filters.append("b.allotment_date <= :to_date")
        params["to_date"] = to_date

    where_sql = ""
    if filters:
        where_sql = " AND " + " AND ".join(filters)

    rows = db.execute(
        text(
            f"""
            WITH latest_assignment AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    DATE(occurred_at) AS allotment_date
                FROM workflow_events
                WHERE event_type = 'claim_assigned'
                ORDER BY claim_id, occurred_at DESC
            ),
            latest_report AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    export_uri
                FROM report_versions
                ORDER BY claim_id, version_no DESC
            ),
            doctor_saved_reports AS (
                SELECT
                    claim_id,
                    1 AS has_doctor_saved
                FROM report_versions
                WHERE NULLIF(TRIM(COALESCE(report_markdown, '')), '') IS NOT NULL
                  AND NOT ({system_report_expr})
                GROUP BY claim_id
            ),
            upload_meta AS (
                SELECT
                    claim_id,
                    report_export_status,
                    tagging,
                    subtagging,
                    opinion
                FROM claim_report_uploads
            ),
            legacy_data AS (
                SELECT
                    claim_id,
                    legacy_payload,
                    updated_at AS legacy_updated_at
                FROM claim_legacy_data
            ),
            base AS (
                SELECT
                    ldata.claim_id,
                    LOWER(TRIM(COALESCE(CAST(c.status AS TEXT), ''))) AS claim_status,
                    CASE WHEN NULLIF(TRIM(COALESCE(c.assigned_doctor_id, '')), '') IS NOT NULL THEN 1 ELSE 0 END AS is_allotted_to_doctor,
                    CASE WHEN COALESCE(dsr.has_doctor_saved, 0) = 1 THEN 1 ELSE 0 END AS has_doctor_saved,
                    COALESCE(
                        CASE
                            WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{{4}}-\d{{2}}-\d{{2}}$'
                                THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'YYYY-MM-DD')
                            WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{{4}}-\d{{2}}-\d{{2}}\s+\d{{2}}:\d{{2}}:\d{{2}}$'
                                THEN TO_TIMESTAMP(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'YYYY-MM-DD HH24:MI:SS')::date
                            WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{{2}}-\d{{2}}-\d{{4}}$'
                                THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD-MM-YYYY')
                            WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{{2}}/\d{{2}}/\d{{4}}$'
                                THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD/MM/YYYY')
                            ELSE NULL
                        END,
                        DATE(ldata.legacy_updated_at),
                        la.allotment_date,
                        DATE(c.updated_at)
                    ) AS allotment_date,
                    CASE
                        WHEN NULLIF(TRIM(COALESCE(um.tagging, '')), '') IS NOT NULL
                          OR NULLIF(TRIM(COALESCE(um.subtagging, '')), '') IS NOT NULL
                          OR NULLIF(TRIM(COALESCE(um.opinion, '')), '') IS NOT NULL
                          OR LOWER(TRIM(COALESCE(um.report_export_status, 'pending'))) = 'uploaded'
                          OR COALESCE(rv.export_uri, '') <> ''
                        THEN 1
                        ELSE 0
                    END AS is_uploaded
                FROM legacy_data ldata
                LEFT JOIN claims c ON c.id = ldata.claim_id
                LEFT JOIN latest_assignment la ON la.claim_id = ldata.claim_id
                LEFT JOIN upload_meta um ON um.claim_id = ldata.claim_id
                LEFT JOIN latest_report rv ON rv.claim_id = ldata.claim_id
                LEFT JOIN doctor_saved_reports dsr ON dsr.claim_id = ldata.claim_id
            )
            SELECT
                b.allotment_date,
                COUNT(*) FILTER (WHERE b.claim_status = 'completed' AND b.is_uploaded = 1) AS completed_count,
                COUNT(*) FILTER (WHERE b.is_allotted_to_doctor = 1 AND b.has_doctor_saved = 0) AS pending_count,
                COUNT(*) FILTER (WHERE b.claim_status = 'completed' AND b.is_uploaded = 1) AS uploaded_count,
                COUNT(*) AS total_count
            FROM base b
            WHERE b.allotment_date IS NOT NULL
            {where_sql}
            GROUP BY b.allotment_date
            ORDER BY b.allotment_date DESC
            """
        ),
        params,
    ).mappings().all()

    return {
        "items": [
            {
                "allotment_date": str(r["allotment_date"]),
                "assigned_count": int(r["total_count"] or 0),
                "pending_count": int(r["pending_count"] or 0),
                "completed_count": int(r["completed_count"] or 0),
                "uploaded_count": int(r["uploaded_count"] or 0),
            }
            for r in rows
        ]
    }


@router.get("/allotment-date-wise/claims")
def allotment_date_wise_claims(
    bucket: str = Query(default="all"),
    allotment_date: str | None = Query(default=None),
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
    limit: int = Query(default=5000, ge=1, le=5000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    _current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
) -> dict:
    _ensure_claim_legacy_data_table(db)
    _ensure_claim_report_uploads_table(db)
    system_report_expr = _system_report_sql("created_by")

    normalized_bucket = str(bucket or "all").strip().lower()
    if normalized_bucket not in {"all", "pending", "completed"}:
        normalized_bucket = "all"

    filters: list[str] = ["b.allotment_date IS NOT NULL"]
    params: dict[str, Any] = {"limit": limit, "offset": offset}

    if _is_valid_date(allotment_date):
        filters.append("b.allotment_date = :allotment_date")
        params["allotment_date"] = allotment_date
    if _is_valid_date(from_date):
        filters.append("b.allotment_date >= :from_date")
        params["from_date"] = from_date
    if _is_valid_date(to_date):
        filters.append("b.allotment_date <= :to_date")
        params["to_date"] = to_date

    if normalized_bucket == "completed":
        filters.append("(b.claim_status = 'completed' AND b.is_uploaded = 1)")
    elif normalized_bucket == "pending":
        filters.append("b.is_allotted_to_doctor = 1 AND b.has_doctor_saved = 0")

    where_sql = "WHERE " + " AND ".join(filters)

    cte_sql = f"""
        WITH latest_assignment AS (
            SELECT DISTINCT ON (claim_id)
                claim_id,
                DATE(occurred_at) AS allotment_date
            FROM workflow_events
            WHERE event_type = 'claim_assigned'
            ORDER BY claim_id, occurred_at DESC
        ),
        latest_report AS (
            SELECT DISTINCT ON (claim_id)
                claim_id,
                export_uri
            FROM report_versions
            ORDER BY claim_id, version_no DESC
        ),
        doctor_saved_reports AS (
            SELECT
                claim_id,
                1 AS has_doctor_saved
            FROM report_versions
            WHERE NULLIF(TRIM(COALESCE(report_markdown, '')), '') IS NOT NULL
              AND NOT ({system_report_expr})
            GROUP BY claim_id
        ),
        upload_meta AS (
            SELECT
                claim_id,
                report_export_status,
                tagging,
                subtagging,
                opinion
            FROM claim_report_uploads
        ),
        legacy_data AS (
            SELECT
                claim_id,
                legacy_payload,
                updated_at AS legacy_updated_at
            FROM claim_legacy_data
        ),
        base AS (
            SELECT
                ldata.claim_id,
                COALESCE(c.external_claim_id, '') AS external_claim_id,
                COALESCE(c.patient_name, '') AS patient_name,
                COALESCE(c.assigned_doctor_id, '') AS assigned_doctor_id,
                LOWER(TRIM(COALESCE(CAST(c.status AS TEXT), ''))) AS claim_status,
                CASE WHEN NULLIF(TRIM(COALESCE(c.assigned_doctor_id, '')), '') IS NOT NULL THEN 1 ELSE 0 END AS is_allotted_to_doctor,
                CASE WHEN COALESCE(dsr.has_doctor_saved, 0) = 1 THEN 1 ELSE 0 END AS has_doctor_saved,
                COALESCE(
                    CASE
                        WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\\d{{4}}-\\d{{2}}-\\d{{2}}$'
                            THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'YYYY-MM-DD')
                        WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\\d{{4}}-\\d{{2}}-\\d{{2}}\\s+\\d{{2}}:\\d{{2}}:\\d{{2}}$'
                            THEN TO_TIMESTAMP(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'YYYY-MM-DD HH24:MI:SS')::date
                        WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\\d{{2}}-\\d{{2}}-\\d{{4}}$'
                            THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD-MM-YYYY')
                        WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\\d{{2}}/\\d{{2}}/\\d{{4}}$'
                            THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD/MM/YYYY')
                        ELSE NULL
                    END,
                    DATE(ldata.legacy_updated_at),
                    la.allotment_date,
                    DATE(c.updated_at)
                ) AS allotment_date,
                CASE
                    WHEN NULLIF(TRIM(COALESCE(um.tagging, '')), '') IS NOT NULL
                      OR NULLIF(TRIM(COALESCE(um.subtagging, '')), '') IS NOT NULL
                      OR NULLIF(TRIM(COALESCE(um.opinion, '')), '') IS NOT NULL
                      OR LOWER(TRIM(COALESCE(um.report_export_status, 'pending'))) = 'uploaded'
                      OR COALESCE(rv.export_uri, '') <> ''
                    THEN 1
                    ELSE 0
                END AS is_uploaded
            FROM legacy_data ldata
            LEFT JOIN claims c ON c.id = ldata.claim_id
            LEFT JOIN latest_assignment la ON la.claim_id = ldata.claim_id
            LEFT JOIN upload_meta um ON um.claim_id = ldata.claim_id
            LEFT JOIN latest_report rv ON rv.claim_id = ldata.claim_id
            LEFT JOIN doctor_saved_reports dsr ON dsr.claim_id = ldata.claim_id
        )
    """

    total = db.execute(
        text(
            f"""
            {cte_sql}
            SELECT COUNT(*)
            FROM base b
            {where_sql}
            """
        ),
        params,
    ).scalar_one()

    rows = db.execute(
        text(
            f"""
            {cte_sql}
            SELECT
                b.claim_id,
                b.external_claim_id,
                b.patient_name,
                b.assigned_doctor_id,
                b.claim_status,
                b.allotment_date,
                CASE
                    WHEN b.is_allotted_to_doctor = 1 AND b.has_doctor_saved = 0 THEN 'pending'
                    WHEN b.claim_status = 'completed' AND b.is_uploaded = 1 THEN 'completed'
                    ELSE 'other'
                END AS bucket
            FROM base b
            {where_sql}
            ORDER BY b.allotment_date DESC, b.external_claim_id ASC
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).mappings().all()

    items = [
        {
            "claim_uuid": str(r.get("claim_id") or ""),
            "external_claim_id": str(r.get("external_claim_id") or ""),
            "patient_name": str(r.get("patient_name") or ""),
            "assigned_doctor_id": str(r.get("assigned_doctor_id") or ""),
            "status": str(r.get("claim_status") or ""),
            "allotment_date": str(r.get("allotment_date") or ""),
            "bucket": str(r.get("bucket") or "pending"),
        }
        for r in rows
    ]

    return {
        "total": int(total or 0),
        "bucket": normalized_bucket,
        "items": items,
    }
@router.get("/dashboard-overview")
def dashboard_overview(
    db: Session = Depends(get_db),
    _current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
) -> dict:
    _ensure_claim_report_uploads_table(db)
    _ensure_claim_legacy_data_table(db)

    day_rows = db.execute(
        text(
            """
            WITH latest_assignment AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    DATE(occurred_at) AS allotment_date
                FROM workflow_events
                WHERE event_type = 'claim_assigned'
                ORDER BY claim_id, occurred_at DESC
            ),
            legacy_data AS (
                SELECT claim_id, legacy_payload
                FROM claim_legacy_data
            ),
            completed_base AS (
                SELECT
                    c.status,
                    COALESCE(
                        la.allotment_date,
                        CASE
                            WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{4}-\d{2}-\d{2}$'
                                THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'YYYY-MM-DD')
                            WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{2}-\d{2}-\d{4}$'
                                THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD-MM-YYYY')
                            WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{2}/\d{2}/\d{4}$'
                                THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD/MM/YYYY')
                            ELSE NULL
                        END,
                        DATE(c.updated_at)
                    ) AS allotment_date
                FROM claims c
                LEFT JOIN latest_assignment la ON la.claim_id = c.id
                LEFT JOIN legacy_data ldata ON ldata.claim_id = c.id
            )
            SELECT
                cb.allotment_date AS completed_date,
                COUNT(*) AS completed_count
            FROM completed_base cb
            WHERE cb.status = 'completed'
              AND cb.allotment_date >= DATE_TRUNC('month', CURRENT_DATE)::date
              AND cb.allotment_date < (DATE_TRUNC('month', CURRENT_DATE) + INTERVAL '1 month')::date
            GROUP BY cb.allotment_date
            ORDER BY cb.allotment_date DESC
            LIMIT 60
            """
        )
    ).mappings().all()

    fraud_row = db.execute(
        text(
            """
            WITH upload_meta AS (
                SELECT
                    claim_id,
                    tagging
                FROM claim_report_uploads
            ),
            legacy_data AS (
                SELECT
                    claim_id,
                    legacy_payload
                FROM claim_legacy_data
            ),
            fraud_base AS (
                SELECT
                    LOWER(
                        TRIM(
                            COALESCE(NULLIF(TRIM(COALESCE(um.tagging, '')), ''), '')
                        )
                    ) AS tagging_value,
                    TRIM(
                        COALESCE(
                            ldata.legacy_payload->>'claim_amount',
                            ldata.legacy_payload->>'claimamount',
                            ldata.legacy_payload->>'claim amount',
                            ''
                        )
                    ) AS claim_amount_text
                FROM claims c
                LEFT JOIN upload_meta um ON um.claim_id = c.id
                LEFT JOIN legacy_data ldata ON ldata.claim_id = c.id
            )
            SELECT
                COUNT(*) AS fraud_tagged_savings_cases,
                COALESCE(
                    SUM(
                        CASE
                            WHEN claim_amount_text ~* '^\\s*(inr|rs\\.?|₹)?\\s*[0-9]{1,3}(,[0-9]{2,3})*(\\.[0-9]+)?\\s*(/-)?\\s*$'
                              OR claim_amount_text ~* '^\\s*(inr|rs\\.?|₹)?\\s*[0-9]+(\\.[0-9]+)?\\s*(/-)?\\s*$'
                            THEN NULLIF(
                                REGEXP_REPLACE(claim_amount_text, '[^0-9.]', '', 'g'),
                                ''
                            )::NUMERIC
                            ELSE 0
                        END
                    ),
                    0
                ) AS fraud_tagged_savings_amount
            FROM fraud_base
            WHERE tagging_value IN ('fraudulent', 'fraudlent')
            """
        )
    ).mappings().one()

    assignee_rows = db.execute(
        text(
            """
            WITH claim_assignees AS (
                SELECT
                    c.id AS claim_id,
                    c.status,
                    assignee AS assignee_key
                FROM claims c
                CROSS JOIN LATERAL unnest(
                    string_to_array(
                        LOWER(REPLACE(COALESCE(c.assigned_doctor_id, ''), ' ', '')),
                        ','
                    )
                ) AS assignee
                WHERE NULLIF(TRIM(COALESCE(c.assigned_doctor_id, '')), '') IS NOT NULL
                  AND NULLIF(TRIM(assignee), '') IS NOT NULL
            ),
            assignee_stats AS (
                SELECT
                    ca.assignee_key,
                    COUNT(*) FILTER (WHERE ca.status = 'completed') AS completed_count,
                    COUNT(*) FILTER (WHERE ca.status NOT IN ('completed', 'withdrawn')) AS pending_count
                FROM claim_assignees ca
                GROUP BY ca.assignee_key
            )
            SELECT
                COALESCE(u.username, s.assignee_key) AS username,
                COALESCE(CAST(u.role AS TEXT), '') AS role,
                CAST(s.completed_count AS INTEGER) AS completed_count,
                CAST(s.pending_count AS INTEGER) AS pending_count,
                CAST(s.completed_count + s.pending_count AS INTEGER) AS total_count
            FROM assignee_stats s
            LEFT JOIN users u
                ON LOWER(u.username) = s.assignee_key
            ORDER BY (s.completed_count + s.pending_count) DESC, COALESCE(u.username, s.assignee_key) ASC
            LIMIT 500
            """
        )
    ).mappings().all()

    return {
        "day_wise_completed": [
            {
                "date": str(r.get("completed_date") or ""),
                "completed": int(r.get("completed_count") or 0),
            }
            for r in day_rows
        ],
        "assignee_wise": [
            {
                "username": str(r.get("username") or "-"),
                "role": str(r.get("role") or ""),
                "completed": int(r.get("completed_count") or 0),
                "pending": int(r.get("pending_count") or 0),
                "total": int(r.get("total_count") or 0),
            }
            for r in assignee_rows
        ],
        "fraud_tagged_savings_cases": int(fraud_row.get("fraud_tagged_savings_cases") or 0),
        "fraud_tagged_savings_amount": float(fraud_row.get("fraud_tagged_savings_amount") or 0),
    }
@router.get("/doctor-completion-stats")
def doctor_completion_stats(
    month: str | None = Query(default=None),
    doctor_username: str | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user, UserRole.doctor)),
) -> dict:
    _ensure_claim_completed_at_column(db)
    month_text = str(month or "").strip()
    selected_month_start: date | None = None
    if month_text:
        if not re.fullmatch(r"\d{4}-\d{2}", month_text):
            raise HTTPException(status_code=400, detail="month must be in YYYY-MM format.")
        try:
            selected_month_start = date.fromisoformat(f"{month_text}-01")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid month value.") from exc

    if current_user.role == UserRole.doctor:
        scoped_doctor_key = _normalize_doctor_key(current_user.username)
        scoped_doctor_label = str(current_user.username or "").strip()
    else:
        requested_doctor = _normalize_optional_text(doctor_username)
        scoped_doctor_key = _normalize_doctor_key(requested_doctor)
        scoped_doctor_label = requested_doctor

    month_rows = db.execute(
        text(
            """
            WITH completed_claims AS (
                SELECT
                    c.id,
                    DATE(COALESCE(c.completed_at, c.updated_at) AT TIME ZONE 'Asia/Kolkata') AS completed_date,
                    COALESCE(c.assigned_doctor_id, '') AS assigned_doctor_id
                FROM claims c
                WHERE c.status = 'completed'
                  AND COALESCE(c.completed_at, c.updated_at) IS NOT NULL
            ),
            scoped_claims AS (
                SELECT
                    cc.id,
                    cc.completed_date
                FROM completed_claims cc
                WHERE :doctor_key = ''
                   OR EXISTS (
                        SELECT 1
                        FROM unnest(
                            string_to_array(
                                regexp_replace(LOWER(cc.assigned_doctor_id), '[^a-z0-9,]+', '', 'g'),
                                ','
                            )
                        ) AS token
                        WHERE NULLIF(token, '') IS NOT NULL
                          AND token = :doctor_key
                   )
            )
            SELECT
                DATE_TRUNC('month', sc.completed_date)::date AS month_start,
                TO_CHAR(DATE_TRUNC('month', sc.completed_date), 'YYYY-MM') AS month_key,
                TO_CHAR(DATE_TRUNC('month', sc.completed_date), 'Mon YYYY') AS month_label,
                COUNT(*)::integer AS closed_count
            FROM scoped_claims sc
            GROUP BY DATE_TRUNC('month', sc.completed_date)
            ORDER BY DATE_TRUNC('month', sc.completed_date) DESC
            LIMIT 36
            """
        ),
        {"doctor_key": scoped_doctor_key},
    ).mappings().all()

    if selected_month_start is None and month_rows:
        top_row = month_rows[0]
        top_month = top_row.get("month_start")
        if isinstance(top_month, date):
            selected_month_start = top_month
        else:
            top_key = str(top_row.get("month_key") or "").strip()
            if re.fullmatch(r"\d{4}-\d{2}", top_key):
                selected_month_start = date.fromisoformat(f"{top_key}-01")

    day_rows: list[dict[str, Any]] = []
    if selected_month_start is not None:
        day_rows = db.execute(
            text(
                """
                WITH completed_claims AS (
                    SELECT
                        c.id,
                        DATE(COALESCE(c.completed_at, c.updated_at) AT TIME ZONE 'Asia/Kolkata') AS completed_date,
                        COALESCE(c.assigned_doctor_id, '') AS assigned_doctor_id
                    FROM claims c
                    WHERE c.status = 'completed'
                      AND COALESCE(c.completed_at, c.updated_at) IS NOT NULL
                ),
                scoped_claims AS (
                    SELECT
                        cc.id,
                        cc.completed_date
                    FROM completed_claims cc
                    WHERE :doctor_key = ''
                       OR EXISTS (
                            SELECT 1
                            FROM unnest(
                                string_to_array(
                                    regexp_replace(LOWER(cc.assigned_doctor_id), '[^a-z0-9,]+', '', 'g'),
                                    ','
                                )
                            ) AS token
                            WHERE NULLIF(token, '') IS NOT NULL
                              AND token = :doctor_key
                       )
                )
                SELECT
                    sc.completed_date AS completed_date,
                    COUNT(*)::integer AS closed_count
                FROM scoped_claims sc
                WHERE sc.completed_date >= :month_start
                  AND sc.completed_date < (:month_start + INTERVAL '1 month')::date
                GROUP BY sc.completed_date
                ORDER BY sc.completed_date DESC
                """
            ),
            {"doctor_key": scoped_doctor_key, "month_start": selected_month_start},
        ).mappings().all()

    selected_month_value = selected_month_start.strftime("%Y-%m") if selected_month_start else ""
    return {
        "doctor_scope": scoped_doctor_label or "all",
        "selected_month": selected_month_value,
        "month_wise_closed": [
            {
                "month": str(r.get("month_key") or ""),
                "label": str(r.get("month_label") or r.get("month_key") or ""),
                "closed": int(r.get("closed_count") or 0),
            }
            for r in month_rows
        ],
        "day_wise_closed": [
            {
                "date": str(r.get("completed_date") or ""),
                "closed": int(r.get("closed_count") or 0),
            }
            for r in day_rows
        ],
    }
@router.get("/claim-document-status")
def claim_document_status(
    search_claim: str | None = Query(default=None),
    allotment_date: str | None = Query(default=None),
    status_filter: str = Query(default="all"),
    doctor_filter: str | None = Query(default=None),
    document_upload: str = Query(default="all"),
    exclude_tagged: bool = Query(default=False),
    exclude_completed: bool = Query(default=False),
    exclude_completed_uploaded: bool = Query(default=False),
    exclude_withdrawn: bool = Query(default=False),
    sort_order: str = Query(default="desc"),
    limit: int = Query(default=200, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user, UserRole.doctor, UserRole.auditor)),
) -> dict:
    _ensure_claim_legacy_data_table(db)
    _ensure_claim_report_uploads_table(db)

    valid_statuses = {
        "ready_for_assignment",
        "waiting_for_documents",
        "pending",
        "in_review",
        "needs_qc",
        "completed",
        "withdrawn",
        "all",
    }
    normalized_status = (status_filter or "all").strip().lower()
    if normalized_status not in valid_statuses:
        normalized_status = "all"

    normalized_document_upload = (document_upload or "all").strip().lower()
    if normalized_document_upload not in {"all", "yes", "no"}:
        normalized_document_upload = "all"

    order_sql = "ASC" if (sort_order or "").strip().lower() == "asc" else "DESC"

    filters: list[str] = []
    params: dict[str, Any] = {"limit": limit, "offset": offset}
    completed_uploaded_expr = (
        "((NULLIF(TRIM(COALESCE(um.tagging, '')), '') IS NOT NULL "
        "OR NULLIF(TRIM(COALESCE(um.subtagging, '')), '') IS NOT NULL "
        "OR NULLIF(TRIM(COALESCE(um.opinion, '')), '') IS NOT NULL) "
        "OR COALESCE(um.report_export_status, 'pending') = 'uploaded' "
        "OR COALESCE(rv.export_uri, '') <> '')"
    )

    if search_claim and search_claim.strip():
        filters.append("LOWER(c.external_claim_id) LIKE :search_claim")
        params["search_claim"] = f"%{search_claim.strip().lower()}%"

    if _is_valid_date(allotment_date):
        filters.append(
            """
            COALESCE(
                la.allotment_date,
                CASE
                    WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\\d{4}-\\d{2}-\\d{2}$'
                        THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'YYYY-MM-DD')
                    WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\\d{2}-\\d{2}-\\d{4}$'
                        THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD-MM-YYYY')
                    WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\\d{2}/\\d{2}/\\d{4}$'
                        THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD/MM/YYYY')
                    ELSE NULL
                END
            ) = :allotment_date
            """
        )
        params["allotment_date"] = allotment_date

    if normalized_document_upload == "yes":
        filters.append("COALESCE(ds.documents, 0) > 0")
    elif normalized_document_upload == "no":
        filters.append("COALESCE(ds.documents, 0) = 0")

    if normalized_status == "pending":
        filters.append("c.status = 'waiting_for_documents'")
        filters.append("COALESCE(ds.documents, 0) > 0")
    elif normalized_status == "waiting_for_documents":
        filters.append("c.status = 'waiting_for_documents'")
        filters.append("COALESCE(ds.documents, 0) = 0")
    elif normalized_status != "all":
        filters.append("c.status = :status_filter")
        params["status_filter"] = normalized_status
    if exclude_completed:
        filters.append("c.status <> 'completed'")
    if exclude_completed_uploaded:
        filters.append(f"NOT (c.status = 'completed' AND {completed_uploaded_expr})")
    auto_exclude_withdrawn = normalized_status != "withdrawn"
    if exclude_withdrawn or auto_exclude_withdrawn:
        filters.append("c.status <> 'withdrawn'")
    if exclude_tagged:
        filters.append("NULLIF(TRIM(COALESCE(um.tagging, '')), '') IS NULL")
    effective_doctors = _split_doctor_filter(doctor_filter)
    if current_user.role == UserRole.doctor:
        doctor_token = _normalize_doctor_token(current_user.username)
        effective_doctors = [doctor_token] if doctor_token else []

    if effective_doctors:
        doctor_clauses: list[str] = []
        for idx, doctor in enumerate(effective_doctors):
            key = f"doctor_{idx}"
            params[key] = doctor
            doctor_clauses.append(
                f":{key} = ANY(string_to_array(regexp_replace(LOWER(COALESCE(c.assigned_doctor_id, '')), '[^a-z0-9,]+', '', 'g'), ','))"
            )
        filters.append("(" + " OR ".join(doctor_clauses) + ")")

    where_sql = ("WHERE " + " AND ".join(filters)) if filters else ""

    total = db.execute(
        text(
            f"""
            WITH latest_assignment AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    occurred_at AS assigned_at,
                    DATE(occurred_at) AS allotment_date
                FROM workflow_events
                WHERE event_type = 'claim_assigned'
                ORDER BY claim_id, occurred_at DESC
            ),
            doc_stats AS (
                SELECT
                    cd.claim_id,
                    COUNT(*) AS documents,
                    SUM(
                        CASE
                            WHEN (cd.metadata->>'merge_source_file_count') ~ '^[0-9]+$'
                                AND CAST(cd.metadata->>'merge_source_file_count' AS INTEGER) > 0
                                THEN CAST(cd.metadata->>'merge_source_file_count' AS INTEGER)
                            WHEN (cd.metadata->>'merge_accepted_file_count') ~ '^[0-9]+$'
                                AND CAST(cd.metadata->>'merge_accepted_file_count' AS INTEGER) > 0
                                THEN CAST(cd.metadata->>'merge_accepted_file_count' AS INTEGER)
                            ELSE 1
                        END
                    ) AS source_files,
                    MAX(cd.uploaded_at) AS last_upload,
                    (
                        ARRAY_REMOVE(
                            ARRAY_AGG(NULLIF(TRIM(COALESCE(cd.uploaded_by, '')), '') ORDER BY cd.uploaded_at DESC NULLS LAST),
                            NULL
                        )
                    )[1] AS last_uploaded_by
                FROM claim_documents cd
                GROUP BY cd.claim_id
            ),
            latest_report AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    export_uri
                FROM report_versions
                ORDER BY claim_id, version_no DESC
            ),
            upload_meta AS (
                SELECT
                    claim_id,
                    report_export_status,
                    tagging,
                    subtagging,
                    opinion
                FROM claim_report_uploads
            )
            SELECT COUNT(*)
            FROM claims c
            LEFT JOIN latest_assignment la ON la.claim_id = c.id
            LEFT JOIN doc_stats ds ON ds.claim_id = c.id
            LEFT JOIN latest_report rv ON rv.claim_id = c.id
            LEFT JOIN upload_meta um ON um.claim_id = c.id
            {where_sql}
            """
        ),
        params,
    ).scalar_one()

    rows = db.execute(
        text(
            f"""
            WITH latest_assignment AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    occurred_at AS assigned_at,
                    DATE(occurred_at) AS allotment_date
                FROM workflow_events
                WHERE event_type = 'claim_assigned'
                ORDER BY claim_id, occurred_at DESC
            ),
            doc_stats AS (
                SELECT
                    cd.claim_id,
                    COUNT(*) AS documents,
                    SUM(
                        CASE
                            WHEN (cd.metadata->>'merge_source_file_count') ~ '^[0-9]+$'
                                AND CAST(cd.metadata->>'merge_source_file_count' AS INTEGER) > 0
                                THEN CAST(cd.metadata->>'merge_source_file_count' AS INTEGER)
                            WHEN (cd.metadata->>'merge_accepted_file_count') ~ '^[0-9]+$'
                                AND CAST(cd.metadata->>'merge_accepted_file_count' AS INTEGER) > 0
                                THEN CAST(cd.metadata->>'merge_accepted_file_count' AS INTEGER)
                            ELSE 1
                        END
                    ) AS source_files,
                    MAX(cd.uploaded_at) AS last_upload,
                    (
                        ARRAY_REMOVE(
                            ARRAY_AGG(NULLIF(TRIM(COALESCE(cd.uploaded_by, '')), '') ORDER BY cd.uploaded_at DESC NULLS LAST),
                            NULL
                        )
                    )[1] AS last_uploaded_by
                FROM claim_documents cd
                GROUP BY cd.claim_id
            ),
            latest_report AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    export_uri
                FROM report_versions
                ORDER BY claim_id, version_no DESC
            ),
            upload_meta AS (
                SELECT
                    claim_id,
                    report_export_status,
                    tagging,
                    subtagging,
                    opinion
                FROM claim_report_uploads
            ),
            latest_decision AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    recommendation,
                    explanation_summary,
                    generated_at
                FROM decision_results
                WHERE is_active = TRUE
                ORDER BY claim_id, generated_at DESC
            ),
            latest_auditor_learning AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    NULLIF(TRIM(COALESCE(notes, '')), '') AS learning_note
                FROM feedback_labels
                WHERE LOWER(TRIM(label_type)) = 'auditor_report_learning'
                ORDER BY claim_id, created_at DESC
            ),
            legacy_data AS (
                SELECT claim_id, legacy_payload
                FROM claim_legacy_data
            )
            SELECT
                c.id,
                c.external_claim_id,
                c.assigned_doctor_id,
                c.tags,
                c.status,
                CASE
                    WHEN c.status = 'waiting_for_documents' AND COALESCE(ds.documents, 0) > 0 THEN 'pending'
                    ELSE c.status::text
                END AS status_display,
                la.assigned_at,
                la.allotment_date,
                COALESCE(ds.documents, 0) AS documents,
                COALESCE(ds.source_files, 0) AS source_files,
                ds.last_upload,
                COALESCE(ds.last_uploaded_by, '') AS last_uploaded_by,
                COALESCE(NULLIF(TRIM(ld.explanation_summary), ''), COALESCE(ld.recommendation::text, 'Pending')) AS final_status,
                COALESCE(
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'doa_date', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'doa', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'doa date', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'date_of_admission', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'date of admission', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'admission_date', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'admission date', '')), '')
                ) AS doa_date,
                COALESCE(
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'dod_date', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'dod', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'dod date', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'date_of_discharge', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'date of discharge', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'discharge_date', '')), ''),
                    NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'discharge date', '')), '')
                ) AS dod_date,
                COALESCE(al.learning_note, '') AS auditor_learning,
                ldata.legacy_payload AS legacy_payload
            FROM claims c
            LEFT JOIN latest_assignment la ON la.claim_id = c.id
            LEFT JOIN doc_stats ds ON ds.claim_id = c.id
            LEFT JOIN latest_report rv ON rv.claim_id = c.id
            LEFT JOIN upload_meta um ON um.claim_id = c.id
            LEFT JOIN latest_decision ld ON ld.claim_id = c.id
            LEFT JOIN latest_auditor_learning al ON al.claim_id = c.id
            LEFT JOIN legacy_data ldata ON ldata.claim_id = c.id
            {where_sql}
            ORDER BY COALESCE(la.allotment_date, DATE(c.updated_at)) {order_sql}, c.updated_at {order_sql}
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).mappings().all()

    def _legacy_text(payload_obj: Any, *keys: str) -> str:
        if not isinstance(payload_obj, dict):
            return ""
        for key in keys:
            value = payload_obj.get(key)
            if value is None:
                continue
            text_value = str(value).strip()
            if text_value:
                return text_value
        return ""

    def _tag_at(tags_value: Any, idx: int) -> str:
        if isinstance(tags_value, list) and 0 <= idx < len(tags_value):
            return str(tags_value[idx] or "").strip()
        if isinstance(tags_value, str):
            try:
                parsed = json.loads(tags_value)
            except Exception:
                parsed = None
            if isinstance(parsed, list) and 0 <= idx < len(parsed):
                return str(parsed[idx] or "").strip()
        return ""

    items = []
    for r in rows:
        legacy_payload = r.get("legacy_payload") if isinstance(r.get("legacy_payload"), dict) else {}
        tags_value = r.get("tags")
        claim_type = (
            _legacy_text(legacy_payload, "claim_type", "claim type", "case_type", "case type")
            or _tag_at(tags_value, 0)
        )
        treatment_type = (
            _legacy_text(legacy_payload, "treatment_type", "treatment type", "treatment-type")
            or _tag_at(tags_value, 4)
        )
        items.append(
            {
                "id": str(r["id"]),
                "external_claim_id": str(r.get("external_claim_id") or ""),
                "assigned_doctor_id": str(r.get("assigned_doctor_id") or ""),
                "status": str(r.get("status") or ""),
                "status_display": str(r.get("status_display") or ""),
                "assigned_at": str(r.get("assigned_at") or ""),
                "allotment_date": str(r.get("allotment_date") or ""),
                "documents": int(r.get("documents") or 0),
                "source_files": int(r.get("source_files") or 0),
                "last_upload": str(r.get("last_upload") or ""),
                "last_uploaded_by": str(r.get("last_uploaded_by") or ""),
                "final_status": str(r.get("final_status") or "Pending"),
                "doa_date": str(r.get("doa_date") or ""),
                "dod_date": str(r.get("dod_date") or ""),
                "auditor_learning": str(r.get("auditor_learning") or ""),
                "claim_type": claim_type,
                "treatment_type": treatment_type,
                "legacy_payload": legacy_payload,
            }
        )

    return {"total": int(total or 0), "items": items}
@router.get("/export-full-data")
def export_full_data(
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
    allotment_date: str | None = Query(default=None),
    format: str = Query(default="json"),
    db: Session = Depends(get_db),
    _current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
):
    _ensure_claim_report_uploads_table(db)
    _ensure_claim_legacy_data_table(db)

    def _format_status(raw: str | None) -> str:
        v = str(raw or "").strip().lower()
        if v in {"waiting_for_documents", "ready_for_assignment", "pending"}:
            return "pending"
        return v or "pending"

    def _fmt_date(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%d-%m-%Y")
        if isinstance(v, date):
            return v.strftime("%d-%m-%Y")

        s = str(v).strip()
        if not s:
            return ""

        normalized = s.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(normalized)
            return dt.strftime("%d-%m-%Y")
        except Exception:
            pass

        date_formats = [
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d-%m-%Y %H:%M",
            "%d-%m-%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
        ]
        for fmt in date_formats:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%d-%m-%Y")
            except Exception:
                continue

        m = re.search(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", s)
        if m:
            token = m.group(0)
            for fmt in ("%d-%m-%Y", "%d/%m/%Y"):
                try:
                    return datetime.strptime(token, fmt).strftime("%d-%m-%Y")
                except Exception:
                    continue

        m = re.search(r"\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b", s)
        if m:
            token = m.group(0)
            for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
                try:
                    return datetime.strptime(token, fmt).strftime("%d-%m-%Y")
                except Exception:
                    continue

        return s

    def _fmt_datetime(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%d-%m-%Y %H:%M:%S")
        if isinstance(v, date):
            return datetime(v.year, v.month, v.day).strftime("%d-%m-%Y %H:%M:%S")

        s = str(v).strip()
        if not s:
            return ""

        normalized = s.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(normalized)
            return dt.strftime("%d-%m-%Y %H:%M:%S")
        except Exception:
            pass

        datetime_formats = [
            "%d-%m-%Y %H:%M:%S",
            "%d-%m-%Y %H:%M",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%Y/%m/%d",
        ]
        for fmt in datetime_formats:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%d-%m-%Y %H:%M:%S")
            except Exception:
                continue

        return s

    def _tag_at(tags: Any, idx: int) -> str:
        if isinstance(tags, list):
            if 0 <= idx < len(tags):
                return str(tags[idx] or "").strip()
            return ""
        return ""

    def _legacy_get(payload_obj: Any, *keys: str) -> str:
        if not isinstance(payload_obj, dict):
            return ""
        for key in keys:
            value = payload_obj.get(key)
            if value is None:
                continue
            text_value = str(value).strip()
            if text_value:
                return text_value
        return ""

    filters = []
    params: dict[str, Any] = {}

    if _is_valid_date(from_date):
        filters.append("DATE(c.created_at) >= :from_date")
        params["from_date"] = from_date
    if _is_valid_date(to_date):
        filters.append("b.allotment_date <= :to_date")
        params["to_date"] = to_date
    if _is_valid_date(allotment_date):
        filters.append("(EXISTS (SELECT 1 FROM workflow_events we WHERE we.claim_id = c.id AND we.event_type = 'claim_assigned' AND DATE(we.occurred_at) = :allotment_date) OR CASE WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{4}-\d{2}-\d{2}$' THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'YYYY-MM-DD') WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{2}-\d{2}-\d{4}$' THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD-MM-YYYY') WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{2}/\d{2}/\d{4}$' THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD/MM/YYYY') ELSE NULL END = :allotment_date)")
        params["allotment_date"] = allotment_date

    where_sql = ("WHERE " + " AND ".join(filters)) if filters else ""

    rows = db.execute(
        text(
            f"""
            WITH latest_report AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id, report_status, export_uri, version_no, created_at AS report_created_at
                FROM report_versions
                ORDER BY claim_id, version_no DESC
            ),
            latest_assignment AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    DATE(occurred_at) AS allotment_date
                FROM workflow_events
                WHERE event_type = 'claim_assigned'
                ORDER BY claim_id, occurred_at DESC
            ),
            doc_stats AS (
                SELECT claim_id, COUNT(*) AS documents
                FROM claim_documents
                GROUP BY claim_id
            ),
            latest_decision AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    explanation_summary
                FROM decision_results
                WHERE is_active = TRUE
                ORDER BY claim_id, generated_at DESC
            ),
            upload_meta AS (
                SELECT
                    claim_id,
                    report_export_status,
                    tagging,
                    subtagging,
                    opinion,
                    qc_status,
                    updated_by
                FROM claim_report_uploads
            ),
            legacy_data AS (
                SELECT claim_id, legacy_payload
                FROM claim_legacy_data
            )
            SELECT
                ROW_NUMBER() OVER (ORDER BY c.created_at DESC) AS row_id,
                c.external_claim_id,
                c.patient_name,
                c.patient_identifier,
                c.status,
                c.assigned_doctor_id,
                c.priority,
                c.source_channel,
                c.tags,
                c.created_at,
                c.updated_at,
                la.allotment_date,
                COALESCE(ds.documents, 0) AS documents,
                COALESCE(ldec.explanation_summary, '') AS trigger_remarks,
                COALESCE(rv.report_status, 'pending') AS report_status,
                COALESCE(rv.export_uri, '') AS export_uri,
                COALESCE(rv.version_no, 0) AS version_no,
                rv.report_created_at,
                COALESCE(um.report_export_status, CASE WHEN COALESCE(rv.export_uri, '') <> '' THEN 'uploaded' ELSE 'pending' END) AS report_export_status,
                COALESCE(um.tagging, '') AS tagging,
                COALESCE(um.subtagging, '') AS subtagging,
                COALESCE(um.opinion, '') AS opinion,
                CASE WHEN LOWER(REPLACE(REPLACE(COALESCE(um.qc_status, 'no'), ' ', '_'), '-', '_')) IN ('yes', 'qc_yes', 'qcyes', 'qc_done', 'done') THEN 'yes' ELSE 'no' END AS qc_status,
                COALESCE(um.updated_by, '') AS uploaded_by_username,
                u.id AS uploaded_by_user_id,
                ldata.legacy_payload
            FROM claims c
            LEFT JOIN latest_report rv ON rv.claim_id = c.id
            LEFT JOIN latest_assignment la ON la.claim_id = c.id
            LEFT JOIN doc_stats ds ON ds.claim_id = c.id
            LEFT JOIN latest_decision ldec ON ldec.claim_id = c.id
            LEFT JOIN upload_meta um ON um.claim_id = c.id
            LEFT JOIN users u ON u.username = um.updated_by
            LEFT JOIN legacy_data ldata ON ldata.claim_id = c.id
            {where_sql}
            ORDER BY c.created_at DESC
            """
        ),
        params,
    ).mappings().all()

    items = [
        {
            "external_claim_id": str(r.get("external_claim_id") or ""),
            "patient_name": str(r.get("patient_name") or ""),
            "patient_identifier": str(r.get("patient_identifier") or ""),
            "status": _format_status(r.get("status")),
            "assigned_doctor_id": str(r.get("assigned_doctor_id") or ""),
            "priority": int(r.get("priority") or 0),
            "source_channel": str(r.get("source_channel") or ""),
            "created_at": str(r.get("created_at") or ""),
            "updated_at": str(r.get("updated_at") or ""),
            "allotment_date": str(r.get("allotment_date") or ""),
            "report_status": str(r.get("report_status") or ""),
            "export_uri": str(r.get("export_uri") or ""),
            "report_version": int(r.get("version_no") or 0),
            "report_created_at": str(r.get("report_created_at") or ""),
        }
        for r in rows
    ]

    legacy_fieldnames = [
        "claim_date",
        "claim_id",
        "claim_type",
        "policy_number",
        "policy_type",
        "policy_start_date",
        "policy_end_date",
        "benef_name",
        "benef_age",
        "benef_gender",
        "pri_benef_name",
        "benef_sum_insured",
        "relation_type",
        "hospital_name",
        "hospital_pincode",
        "hospital_city",
        "hospital_state",
        "claim_amount",
        "doa_date",
        "dod_date",
        "claimant_ir",
        "hospital_is_network",
        "trigger_remarks",
        "document_required",
        "primary_icd_group",
        "primary_ailment_code",
        "treatment_type",
        "bill_deduction_reason",
        "vendor_name",
        "allocation_date",
        "document_status",
        "final_status",
        "report_export_status",
        "tagging",
        "subtagging",
        "opinion",
        "qc_status",
        "created_at",
        "updated_at",
    ]

    def _claim_id_number(value: Any) -> int | str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        digits = re.sub(r"[^0-9]", "", raw)
        if not digits:
            return raw
        try:
            return int(digits)
        except Exception:
            return raw

    legacy_items: list[dict[str, Any]] = []
    for r in rows:
        tags = r.get("tags")
        status_value = _format_status(r.get("status"))
        payload_obj = r.get("legacy_payload")
        if isinstance(payload_obj, str):
            try:
                payload_obj = json.loads(payload_obj)
            except Exception:
                payload_obj = {}
        if not isinstance(payload_obj, dict):
            payload_obj = {}


        legacy_items.append(
            {
                "claim_date": _fmt_date(_legacy_get(payload_obj, "claim_date") or r.get("created_at")),
                "claim_id": _claim_id_number(_legacy_get(payload_obj, "claim_id") or str(r.get("external_claim_id") or "")), 
                "claim_type": _legacy_get(payload_obj, "claim_type") or _tag_at(tags, 0),
                "policy_number": _legacy_get(payload_obj, "policy_number") or str(r.get("patient_identifier") or ""),
                "policy_type": _legacy_get(payload_obj, "policy_type") or _tag_at(tags, 1),
                "policy_start_date": _fmt_date(_legacy_get(payload_obj, "policy_start_date")),
                "policy_end_date": _fmt_date(_legacy_get(payload_obj, "policy_end_date")),
                "benef_name": _legacy_get(payload_obj, "benef_name") or str(r.get("patient_name") or ""),
                "benef_age": _legacy_get(payload_obj, "benef_age"),
                "benef_gender": _legacy_get(payload_obj, "benef_gender"),
                "pri_benef_name": _legacy_get(payload_obj, "pri_benef_name"),
                "benef_sum_insured": _legacy_get(payload_obj, "benef_sum_insured"),
                "relation_type": _legacy_get(payload_obj, "relation_type"),
                "hospital_name": _legacy_get(payload_obj, "hospital_name") or _tag_at(tags, 3),
                "hospital_pincode": _legacy_get(payload_obj, "hospital_pincode"),
                "hospital_city": _legacy_get(payload_obj, "hospital_city"),
                "hospital_state": _legacy_get(payload_obj, "hospital_state"),
                "claim_amount": _legacy_get(payload_obj, "claim_amount"),
                "doa_date": _fmt_date(_legacy_get(payload_obj, "doa_date", "doa", "doa date", "date_of_admission", "date of admission", "admission_date", "admission date")),
                "dod_date": _fmt_date(_legacy_get(payload_obj, "dod_date", "dod", "dod date", "date_of_discharge", "date of discharge", "discharge_date", "discharge date")),
                "claimant_ir": _legacy_get(payload_obj, "claimant_ir"),
                "hospital_is_network": _legacy_get(payload_obj, "hospital_is_network"),
                "trigger_remarks": _legacy_get(payload_obj, "trigger_remarks") or str(r.get("trigger_remarks") or ""),
                "document_required": _legacy_get(payload_obj, "document_required"),
                "primary_icd_group": _legacy_get(payload_obj, "primary_icd_group") or _tag_at(tags, 2),
                "primary_ailment_code": _legacy_get(payload_obj, "primary_ailment_code"),
                "treatment_type": _legacy_get(payload_obj, "treatment_type") or _tag_at(tags, 4),
                "bill_deduction_reason": _legacy_get(payload_obj, "bill_deduction_reason"),
                "vendor_name": _legacy_get(payload_obj, "vendor_name") or str(r.get("source_channel") or ""),
                "allocation_date": _fmt_date(_legacy_get(payload_obj, "allocation_date") or r.get("allotment_date")),
                "document_status": _legacy_get(payload_obj, "document_status") or ("uploaded" if int(r.get("documents") or 0) > 0 else "pending"),
                "final_status": _legacy_get(payload_obj, "final_status") or status_value,
                "report_export_status": _legacy_get(payload_obj, "report_export_status") or str(r.get("report_export_status") or "pending"),
                "tagging": _legacy_get(payload_obj, "tagging") or str(r.get("tagging") or ""),
                "subtagging": _legacy_get(payload_obj, "subtagging") or str(r.get("subtagging") or ""),
                "opinion": _legacy_get(payload_obj, "opinion") or str(r.get("opinion") or ""),
                "qc_status": _legacy_get(payload_obj, "qc_status") or str(r.get("qc_status") or "no"),
                "created_at": _fmt_datetime(_legacy_get(payload_obj, "created_at") or r.get("created_at")),
                "updated_at": _fmt_datetime(_legacy_get(payload_obj, "updated_at") or r.get("updated_at")),
            }
        )

    normalized_format = format.lower()
    if normalized_format == "csv":
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=legacy_fieldnames, extrasaction="ignore")
        writer.writeheader()
        for item in legacy_items:
            writer.writerow(item)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return PlainTextResponse(
            content=buf.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=user_full_data_{stamp}.csv"},
        )

    if normalized_format in {"excel", "xlsx"}:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            from openpyxl import Workbook
            from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

            def _xlsx_value(v: Any) -> Any:
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    return v
                raw = "" if v is None else str(v)
                cleaned = ILLEGAL_CHARACTERS_RE.sub("", raw)
                # XLSX cell text max is 32767 characters.
                if len(cleaned) > 32767:
                    cleaned = cleaned[:32767]
                return cleaned

            wb = Workbook()
            ws = wb.active
            ws.title = "user_full_data"
            ws.append(legacy_fieldnames)
            for item in legacy_items:
                ws.append([_xlsx_value(item.get(col, "")) for col in legacy_fieldnames])

            out = io.BytesIO()
            wb.save(out)
            return Response(
                content=out.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=user_full_data_{stamp}.xlsx"},
            )
        except Exception:
            buf = io.StringIO()
            writer = csv.DictWriter(buf, fieldnames=legacy_fieldnames, extrasaction="ignore")
            writer.writeheader()
            for item in legacy_items:
                writer.writerow(item)

            return PlainTextResponse(
                content=buf.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=user_full_data_{stamp}.csv"},
            )

    return {"total": len(items), "items": items}




def _ensure_user_bank_details_table(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS user_bank_details (
                id BIGSERIAL PRIMARY KEY,
                user_id BIGINT NOT NULL UNIQUE REFERENCES users(id) ON DELETE CASCADE,
                account_holder_name VARCHAR(255) NOT NULL DEFAULT '',
                bank_name VARCHAR(255) NOT NULL DEFAULT '',
                branch_name VARCHAR(255) NOT NULL DEFAULT '',
                account_number VARCHAR(64) NOT NULL DEFAULT '',
                payment_rate VARCHAR(64) NOT NULL DEFAULT '',
                ifsc_code VARCHAR(32) NOT NULL DEFAULT '',
                upi_id VARCHAR(255) NOT NULL DEFAULT '',
                notes TEXT NOT NULL DEFAULT '',
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                created_by VARCHAR(100) NOT NULL DEFAULT '',
                updated_by VARCHAR(100) NOT NULL DEFAULT '',
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(text("ALTER TABLE user_bank_details ADD COLUMN IF NOT EXISTS payment_rate VARCHAR(64) NOT NULL DEFAULT ''"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_user_bank_details_user_id ON user_bank_details(user_id)"))


def _parse_payment_rate(value: Any) -> float:
    raw = str(value or "").strip()
    if not raw:
        return 0.0
    cleaned = re.sub(r"[^0-9.\-]", "", raw)
    if cleaned in {"", "-", ".", "-."}:
        return 0.0
    try:
        parsed = float(cleaned)
    except Exception:
        return 0.0
    return parsed if parsed >= 0 else 0.0


def _next_month_start(month_start: date) -> date:
    if month_start.month == 12:
        return date(month_start.year + 1, 1, 1)
    return date(month_start.year, month_start.month + 1, 1)


@router.get("/payment-sheet")
def payment_sheet(
    month: str | None = Query(default=None),
    include_zero_cases: bool = Query(default=True),
    db: Session = Depends(get_db),
    _current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin)),
) -> dict:
    _ensure_user_bank_details_table(db)
    _ensure_claim_completed_at_column(db)

    month_text = str(month or "").strip()
    if month_text:
        if not re.fullmatch(r"\d{4}-\d{2}", month_text):
            raise HTTPException(status_code=400, detail="month must be in YYYY-MM format.")
        try:
            month_start = date.fromisoformat(f"{month_text}-01")
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Invalid month value.") from exc
    else:
        ist_now = datetime.now(timezone(timedelta(hours=5, minutes=30))).date()
        this_month_start = date(ist_now.year, ist_now.month, 1)
        month_start = (this_month_start - timedelta(days=1)).replace(day=1)

    month_end = _next_month_start(month_start)

    rows = db.execute(
        text(
            """
            WITH eligible_users AS (
                SELECT
                    u.id AS user_id,
                    u.username,
                    CAST(u.role AS TEXT) AS role,
                    COALESCE(ubd.payment_rate, '') AS payment_rate_raw,
                    COALESCE(ubd.is_active, TRUE) AS bank_is_active
                FROM users u
                LEFT JOIN user_bank_details ubd ON ubd.user_id = u.id
                WHERE CAST(u.role AS TEXT) IN ('super_admin', 'doctor')
            ),
            completed_claim_tokens AS (
                SELECT
                    c.id AS claim_id,
                    token AS doctor_key
                FROM claims c
                CROSS JOIN LATERAL unnest(
                    string_to_array(
                        regexp_replace(LOWER(COALESCE(c.assigned_doctor_id, '')), '[^a-z0-9,]+', '', 'g'),
                        ','
                    )
                ) AS token
                WHERE c.status = 'completed'
                  AND COALESCE(c.completed_at, c.updated_at) IS NOT NULL
                  AND DATE(COALESCE(c.completed_at, c.updated_at) AT TIME ZONE 'Asia/Kolkata') >= :month_start
                  AND DATE(COALESCE(c.completed_at, c.updated_at) AT TIME ZONE 'Asia/Kolkata') < :month_end
                  AND NULLIF(token, '') IS NOT NULL
            ),
            completed_counts AS (
                SELECT
                    ct.doctor_key,
                    COUNT(DISTINCT ct.claim_id)::integer AS completed_cases
                FROM completed_claim_tokens ct
                GROUP BY ct.doctor_key
            )
            SELECT
                eu.user_id,
                eu.username,
                eu.role,
                eu.payment_rate_raw,
                eu.bank_is_active,
                COALESCE(cc.completed_cases, 0)::integer AS completed_cases
            FROM eligible_users eu
            LEFT JOIN completed_counts cc ON LOWER(eu.username) = cc.doctor_key
            ORDER BY LOWER(eu.username) ASC
            """
        ),
        {
            "month_start": month_start,
            "month_end": month_end,
        },
    ).mappings().all()

    items: list[dict[str, Any]] = []
    total_cases = 0
    total_amount = 0.0
    for row in rows:
        completed_cases = int(row.get("completed_cases") or 0)
        if not include_zero_cases and completed_cases <= 0:
            continue
        rate_raw = str(row.get("payment_rate_raw") or "").strip()
        rate_numeric = _parse_payment_rate(rate_raw)
        amount_total = float(rate_numeric * completed_cases)
        total_cases += completed_cases
        total_amount += amount_total
        items.append(
            {
                "user_id": int(row.get("user_id") or 0),
                "username": str(row.get("username") or ""),
                "role": str(row.get("role") or ""),
                "rate_raw": rate_raw,
                "rate_numeric": rate_numeric,
                "completed_cases": completed_cases,
                "amount_total": amount_total,
                "bank_is_active": bool(row.get("bank_is_active")),
            }
        )

    return {
        "month": month_start.strftime("%Y-%m"),
        "month_label": month_start.strftime("%b %Y"),
        "include_zero_cases": bool(include_zero_cases),
        "total_users": len(items),
        "total_cases": int(total_cases),
        "total_amount": float(total_amount),
        "items": items,
    }
