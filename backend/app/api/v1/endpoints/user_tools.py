import logging
import csv
import io
import json
import re
from datetime import date, datetime, timedelta, timezone
from uuid import UUID
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import PlainTextResponse, Response
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.deps.auth import require_roles
from app.core.config import settings
from app.db.session import SessionLocal, get_db
from app.repositories import (
    allotment_reporting_repo,
    claim_legacy_data_repo,
    claim_report_uploads_repo,
    claims_repo,
    dashboard_reporting_repo,
    decision_results_repo,
    model_registry_repo,
    report_versions_repo,
    workflow_events_repo,
)
from app.domain.user_tools.excel_import_use_case import import_claims_from_excel_payload
from app.domain.user_tools.completed_reports_use_case import get_completed_reports
from app.domain.user_tools.completed_report_qc_use_case import (
    CompletedClaimNotFoundError,
    InvalidQcStatusError,
    update_completed_report_qc_status as update_completed_report_qc_status_use_case,
)
from app.domain.user_tools.completed_report_upload_use_case import (
    CompletedClaimNotFoundError as CompletedClaimNotFoundForUploadError,
    InvalidUploadPayloadError,
    update_completed_report_upload_status as update_completed_report_upload_status_use_case,
)
from app.domain.user_tools.completed_report_latest_html_use_case import (
    ClaimNotFoundError as CompletedLatestHtmlClaimNotFoundError,
    ForbiddenError as CompletedLatestHtmlForbiddenError,
    InvalidSourceError as CompletedLatestHtmlInvalidSourceError,
    ReportNotFoundError as CompletedLatestHtmlNotFoundError,
    get_completed_report_latest_html as get_completed_report_latest_html_use_case,
)
from app.domain.user_tools.doctor_completion_stats_use_case import (
    InvalidMonthError,
    get_doctor_completion_stats,
)
from app.domain.user_tools.claim_document_status_use_case import get_claim_document_status
from app.domain.user_tools.payment_sheet_use_case import (
    InvalidMonthError as PaymentSheetInvalidMonthError,
    get_payment_sheet as get_payment_sheet_use_case,
)
from app.domain.user_tools.export_full_data_use_case import (
    ExportBinaryResult,
    export_full_data as export_full_data_use_case,
)
from app.schemas.auth import UserRole
from app.schemas.qc_tools import (
    CompletedReportLatestHtmlResponse,
    CompletedReportQcStatusRequest,
    CompletedReportQcStatusResponse,
    CompletedReportUploadStatusRequest,
    CompletedReportUploadStatusResponse,
    ExcelImportResponse,
)
from app.domain.auth.service import AuthenticatedUser
from app.ml import (
    AUDITOR_QC_LABEL_TYPE,
    MODEL_KEY,
    ensure_model,
    upsert_feedback_label,
)
from app.infrastructure.parsers.sql_dump_parser import iter_table_rows_from_sql_dump_bytes
from app.dependencies.access_control import doctor_matches_assignment

router = APIRouter(prefix="/user-tools", tags=["user-tools"])

logger = logging.getLogger(__name__)
_EMPTY_LIKE_TEXT_VALUES = {
    "na",
    "n/a",
    "not available",
    "none",
    "nil",
    "null",
    "-",
    ".",
}


def _normalize_optional_text(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    if raw.lower() in _EMPTY_LIKE_TEXT_VALUES:
        return ""
    return raw

def _parse_datetime_utc(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        dt = value
    else:
        raw = str(value or "").strip()
        if not raw:
            return None
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except Exception:
            return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _latest_model_trained_recently(db: Session, min_minutes: int) -> bool:
    if min_minutes <= 0:
        return False
    try:
        trained_at_raw = model_registry_repo.get_latest_trained_at(db, model_key=MODEL_KEY)
    except Exception:
        return False
    trained_at = _parse_datetime_utc(trained_at_raw)
    if trained_at is None:
        return False
    cutoff = datetime.now(timezone.utc) - timedelta(minutes=max(0, min_minutes))
    return trained_at >= cutoff


def _run_ml_retrain_background(triggered_by: str, claim_id: str) -> None:
    if not settings.ml_auto_retrain_on_qc_yes:
        return
    db = SessionLocal()
    try:
        min_interval = int(settings.ml_auto_retrain_min_interval_minutes or 0)
        if _latest_model_trained_recently(db, min_interval):
            logger.info(
                "Skipping ML retrain (throttled). trigger=%s claim_id=%s min_interval=%s",
                triggered_by,
                claim_id,
                min_interval,
            )
            return
        model = ensure_model(db=db, force_retrain=True)
        version = str(model.get("version") or "") if isinstance(model, dict) else ""
        examples = int(model.get("num_examples") or 0) if isinstance(model, dict) else 0
        logger.info(
            "ML retrain completed. trigger=%s claim_id=%s version=%s examples=%s",
            triggered_by,
            claim_id,
            version,
            examples,
        )
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        logger.exception(
            "ML retrain failed. trigger=%s claim_id=%s error=%s",
            triggered_by,
            claim_id,
            exc,
        )
    finally:
        db.close()



def _is_valid_date(value: str | None) -> bool:
    if not value:
        return False
    try:
        date.fromisoformat(value)
        return True
    except Exception:
        return False


def _normalize_doctor_key(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _coerce_claim_status(raw: str | None) -> str:
    v = (raw or "").strip().lower()
    valid = {
        "ready_for_assignment",
        "waiting_for_documents",
        "in_review",
        "needs_qc",
        "completed",
        "withdrawn",
    }
    return v if v in valid else "waiting_for_documents"


TAGGING_SUBTAGGING_OPTIONS: dict[str, list[str]] = {
    "Genuine": [
        "Hospitalization verified and found to be genuine",
    ],
    "Fraudulent": [
        "Non cooperation of Hospital / patient during investigation",
        "Circumstantial evidence suggesting of possible fraud",
        "Infalted bills",
        "OPD to IPD conversion",
    ],
}


def _normalize_tagging(value: str | None) -> str:
    normalized = str(value or "").strip().lower()
    if normalized == "genuine":
        return "Genuine"
    if normalized in {"fraudulent", "fraudlent"}:
        return "Fraudulent"
    return ""


def _normalize_subtagging(tagging: str, value: str | None) -> str:
    options = TAGGING_SUBTAGGING_OPTIONS.get(tagging, [])
    raw = str(value or "").strip()
    if not raw:
        return ""
    for option in options:
        if raw.lower() == option.lower():
            return option
    return ""



def _normalize_qc_status(value: str | None) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return "no"
    compact = raw.replace("-", "_").replace(" ", "_")
    if compact in {"yes", "qc_yes", "qcyes", "qc_done", "done"}:
        return "yes"
    if compact in {"no", "qc_no", "qcno", "pending", "not_done"}:
        return "no"
    return ""


def _recommendation_to_feedback_label(raw: str | None) -> str | None:
    recommendation = str(raw or "").strip().lower()
    if recommendation in {"approve", "approved", "admissible", "payable"}:
        return "approve"
    if recommendation in {"reject", "rejected", "inadmissible"}:
        return "reject"
    if recommendation in {"need_more_evidence", "query", "manual_review"}:
        return "need_more_evidence"
    return None

def _system_report_sql(column_expr: str) -> str:
    col = f"LOWER(COALESCE({column_expr}, ''))"
    return (
        f"({col} LIKE 'system:%' OR {col} IN "
        "('system', 'system_ml', 'system-ai', 'ml-system', 'checklist_pipeline'))"
    )


def _report_source_from_created_by(created_by: str | None) -> str:
    raw = str(created_by or '').strip().lower()
    if raw.startswith('system:') or raw in {'system', 'system_ml', 'system-ai', 'ml-system', 'checklist_pipeline'}:
        return 'system'
    return 'doctor'


def _ensure_claim_report_uploads_table(db: Session) -> None:
    claim_report_uploads_repo.ensure_claim_report_uploads_table(db)


def _ensure_claim_legacy_data_table(db: Session) -> None:
    claim_legacy_data_repo.ensure_claim_legacy_data_table(db)



def _ensure_claim_completed_at_column(db: Session) -> None:
    claims_repo.ensure_claim_completed_at_column_and_backfill(db)

def _normalize_doctor_token(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _split_doctor_filter(raw: str | None) -> list[str]:
    tokens: list[str] = []
    for part in str(raw or "").split(","):
        normalized = _normalize_doctor_token(part)
        if normalized:
            tokens.append(normalized)
    return tokens

def _normalize_import_key(value: Any) -> str:
    key = str(value or "").strip().lower()
    if not key:
        return ""
    key = key.replace("\ufeff", "")
    key = re.sub(r"[\s\-/]+", "_", key)
    key = re.sub(r"[^a-z0-9_]", "", key)
    key = re.sub(r"_+", "_", key).strip("_")
    return key


def _row_with_normalized_aliases(raw_map: dict[Any, Any]) -> dict[str, str]:
    norm: dict[str, str] = {}
    for key, value in (raw_map or {}).items():
        raw_key = str(key or "").strip().lower()
        if not raw_key:
            continue
        text_value = str(value or "").strip()
        norm[raw_key] = text_value
        normalized_key = _normalize_import_key(raw_key)
        if normalized_key and normalized_key not in norm:
            norm[normalized_key] = text_value
    return norm


def _parse_csv_rows(payload: bytes) -> list[dict[str, str]]:
    text_payload = payload.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text_payload))
    rows: list[dict[str, str]] = []
    for row in reader:
        rows.append(_row_with_normalized_aliases(row))
    return rows


def _parse_xlsx_rows(payload: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed for xlsx import") from exc

    wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(h or "").strip().lower() for h in headers_raw]
    rows: list[dict[str, str]] = []
    for row_vals in rows_iter:
        raw_row: dict[str, str] = {}
        for idx, val in enumerate(row_vals):
            key = headers[idx] if idx < len(headers) else f"col_{idx+1}"
            raw_row[key] = str(val).strip() if val is not None else ""
        rows.append(_row_with_normalized_aliases(raw_row))
    return rows


def _parse_sql_dump_rows(payload: bytes) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for raw_row in iter_table_rows_from_sql_dump_bytes(payload, "excel_case_uploads"):
        prepared_row: dict[str, str] = {}
        for key, value in (raw_row or {}).items():
            raw_key = str(key or "").strip().lower()
            if not raw_key:
                continue
            if value is None:
                prepared_row[raw_key] = ""
            elif isinstance(value, bool):
                prepared_row[raw_key] = "1" if value else "0"
            else:
                prepared_row[raw_key] = str(value).strip()
        norm = _row_with_normalized_aliases(prepared_row)
        if norm:
            rows.append(norm)
    return rows


def _extract_claim_fields(row: dict[str, str]) -> dict[str, Any] | None:
    external_claim_id = (
        row.get("external_claim_id")
        or row.get("claim_id")
        or row.get("claim no")
        or row.get("claim")
        or ""
    ).strip()
    if not external_claim_id:
        return None

    patient_name = (
        row.get("patient_name")
        or row.get("benef_name")
        or row.get("patient")
        or ""
    ).strip() or None

    patient_identifier = (
        row.get("patient_identifier")
        or row.get("policy_number")
        or row.get("policy no")
        or ""
    ).strip() or None

    assigned_doctor_id = (
        row.get("assigned_doctor_id")
        or row.get("doctor_username")
        or row.get("doctor")
        or ""
    ).strip() or None

    status = _coerce_claim_status(row.get("status"))
    if assigned_doctor_id and status == "waiting_for_documents":
        status = "in_review"

    tags: list[str] = []
    for key in ["claim_type", "policy_type", "diagnosis", "hospital_name", "treatment_type"]:
        val = (row.get(key) or "").strip()
        if val:
            tags.append(val)

    return {
        "external_claim_id": external_claim_id,
        "patient_name": patient_name,
        "patient_identifier": patient_identifier,
        "status": status,
        "assigned_doctor_id": assigned_doctor_id,
        "priority": 3,
        "source_channel": "excel_import",
        "tags": tags,
    }


@router.post("/upload-excel", response_model=ExcelImportResponse)
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
) -> ExcelImportResponse:
    payload = await file.read()
    if not payload:
        raise HTTPException(status_code=400, detail="empty file")

    filename = str(file.filename or "")
    total_rows, inserted, updated, skipped = import_claims_from_excel_payload(
        db,
        payload=payload,
        filename=filename,
        uploaded_by_username=str(current_user.username or ""),
    )
    return ExcelImportResponse(total_rows=total_rows, inserted=inserted, updated=updated, skipped=skipped)

@router.get("/completed-reports")
def completed_reports(
    status_filter: str = Query(default="pending"),
    qc_filter: str = Query(default="no"),
    search_claim: str | None = Query(default=None),
    allotment_date: str | None = Query(default=None),
    doctor_filter: str | None = Query(default=None),
    exclude_tagged: bool = Query(default=False),
    sort_order: str = Query(default="updated_desc"),
    limit: int = Query(default=200, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user, UserRole.auditor, UserRole.doctor)),
) -> dict:
    return get_completed_reports(
        db,
        status_filter=status_filter,
        qc_filter=qc_filter,
        search_claim=search_claim,
        allotment_date=allotment_date,
        doctor_filter=doctor_filter,
        exclude_tagged=exclude_tagged,
        sort_order=sort_order,
        limit=limit,
        offset=offset,
    )


@router.post("/completed-reports/{claim_id}/upload-status", response_model=CompletedReportUploadStatusResponse)
def update_completed_report_upload_status(
    claim_id: UUID,
    payload: CompletedReportUploadStatusRequest,
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
) -> CompletedReportUploadStatusResponse:
    try:
        return update_completed_report_upload_status_use_case(
            db,
            claim_id=claim_id,
            report_export_status=payload.report_export_status,
            tagging=payload.tagging,
            subtagging=payload.subtagging,
            opinion=payload.opinion,
            actor_username=current_user.username,
        )
    except InvalidUploadPayloadError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except CompletedClaimNotFoundForUploadError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@router.post("/completed-reports/{claim_id}/qc-status", response_model=CompletedReportQcStatusResponse)
def update_completed_report_qc_status(
    claim_id: UUID,
    payload: CompletedReportQcStatusRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.auditor)),
) -> CompletedReportQcStatusResponse:
    try:
        result = update_completed_report_qc_status_use_case(
            db,
            claim_id=claim_id,
            qc_status=payload.qc_status,
            actor_username=current_user.username,
        )
    except InvalidQcStatusError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except CompletedClaimNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    if settings.ml_auto_retrain_on_qc_yes:
        background_tasks.add_task(_run_ml_retrain_background, current_user.username, str(claim_id))
        logger.info(
            "Queued ML retrain after QC audit update. claim_id=%s actor=%s qc_status=%s",
            str(claim_id),
            current_user.username,
            str(result.qc_status or ""),
        )

    return result


@router.get("/completed-reports/{claim_id}/latest-html", response_model=CompletedReportLatestHtmlResponse)
def get_completed_report_latest_html(
    claim_id: UUID,
    source: str = Query(default="any"),
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user, UserRole.auditor, UserRole.doctor)),
) -> CompletedReportLatestHtmlResponse:
    try:
        return get_completed_report_latest_html_use_case(
            db,
            claim_id=claim_id,
            source=source,
            current_user_role=current_user.role,
            current_username=current_user.username,
        )
    except CompletedLatestHtmlInvalidSourceError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except CompletedLatestHtmlClaimNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except CompletedLatestHtmlForbiddenError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    except CompletedLatestHtmlNotFoundError as exc:
        # Backward-compat: use-case now returns an empty report response when
        # nothing exists, so this is unlikely; keep as 404 if it happens.
        raise HTTPException(status_code=404, detail=str(exc)) from exc

@router.get("/allotment-date-wise")
def allotment_date_wise(
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
    db: Session = Depends(get_db),
    _current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
) -> dict:
    _ensure_claim_legacy_data_table(db)
    _ensure_claim_report_uploads_table(db)
    from_date_filter = from_date if _is_valid_date(from_date) else None
    to_date_filter = to_date if _is_valid_date(to_date) else None

    rows = allotment_reporting_repo.list_allotment_date_wise_summary(
        db,
        from_date=from_date_filter,
        to_date=to_date_filter,
    )

    return {
        "items": [
            {
                "allotment_date": str(r["allotment_date"]),
                "assigned_count": int(r["total_count"] or 0),
                "pending_count": int(r["pending_count"] or 0),
                "completed_count": int(r["completed_count"] or 0),
                "uploaded_count": int(r["uploaded_count"] or 0),
            }
            for r in rows
        ]
    }


@router.get("/allotment-date-wise/claims")
def allotment_date_wise_claims(
    bucket: str = Query(default="all"),
    allotment_date: str | None = Query(default=None),
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
    limit: int = Query(default=5000, ge=1, le=5000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    _current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
) -> dict:
    _ensure_claim_legacy_data_table(db)
    _ensure_claim_report_uploads_table(db)

    normalized_bucket = str(bucket or "all").strip().lower()
    if normalized_bucket not in {"all", "pending", "completed"}:
        normalized_bucket = "all"

    allotment_date_filter = allotment_date if _is_valid_date(allotment_date) else None
    from_date_filter = from_date if _is_valid_date(from_date) else None
    to_date_filter = to_date if _is_valid_date(to_date) else None

    total, rows = allotment_reporting_repo.list_allotment_date_wise_claims(
        db,
        bucket=normalized_bucket,
        allotment_date=allotment_date_filter,
        from_date=from_date_filter,
        to_date=to_date_filter,
        limit=limit,
        offset=offset,
    )

    items = [
        {
            "claim_uuid": str(r.get("claim_id") or ""),
            "external_claim_id": str(r.get("external_claim_id") or ""),
            "patient_name": str(r.get("patient_name") or ""),
            "assigned_doctor_id": str(r.get("assigned_doctor_id") or ""),
            "status": str(r.get("claim_status") or ""),
            "allotment_date": str(r.get("allotment_date") or ""),
            "bucket": str(r.get("bucket") or "pending"),
        }
        for r in rows
    ]

    return {
        "total": int(total or 0),
        "bucket": normalized_bucket,
        "items": items,
    }
@router.get("/dashboard-overview")
def dashboard_overview(
    db: Session = Depends(get_db),
    _current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
) -> dict:
    _ensure_claim_legacy_data_table(db)
    day_rows = dashboard_reporting_repo.list_day_wise_completed_current_month(db)
    assignee_rows = dashboard_reporting_repo.list_assignee_wise_stats(db)

    return {
        "day_wise_completed": [
            {
                "date": str(r.get("completed_date") or ""),
                "completed": int(r.get("completed_count") or 0),
            }
            for r in day_rows
        ],
        "assignee_wise": [
            {
                "username": str(r.get("username") or "-"),
                "role": str(r.get("role") or ""),
                "completed": int(r.get("completed_count") or 0),
                "pending": int(r.get("pending_count") or 0),
                "total": int(r.get("total_count") or 0),
            }
            for r in assignee_rows
        ],
    }
@router.get("/doctor-completion-stats")
def doctor_completion_stats(
    month: str | None = Query(default=None),
    doctor_username: str | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user, UserRole.doctor)),
) -> dict:
    try:
        return get_doctor_completion_stats(
            db,
            month=month,
            doctor_username=doctor_username,
            current_user_role=current_user.role,
            current_username=current_user.username,
        )
    except InvalidMonthError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
@router.get("/claim-document-status")
def claim_document_status(
    search_claim: str | None = Query(default=None),
    allotment_date: str | None = Query(default=None),
    status_filter: str = Query(default="all"),
    doctor_filter: str | None = Query(default=None),
    document_upload: str = Query(default="all"),
    exclude_tagged: bool = Query(default=False),
    exclude_completed: bool = Query(default=False),
    exclude_completed_uploaded: bool = Query(default=False),
    exclude_withdrawn: bool = Query(default=False),
    sort_order: str = Query(default="desc"),
    limit: int = Query(default=200, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user, UserRole.doctor, UserRole.auditor)),
) -> dict:
    return get_claim_document_status(
        db,
        search_claim=search_claim,
        allotment_date=allotment_date,
        status_filter=status_filter,
        doctor_filter=doctor_filter,
        document_upload=document_upload,
        exclude_tagged=exclude_tagged,
        exclude_completed=exclude_completed,
        exclude_completed_uploaded=exclude_completed_uploaded,
        exclude_withdrawn=exclude_withdrawn,
        sort_order=sort_order,
        limit=limit,
        offset=offset,
        current_user_role=current_user.role,
        current_username=current_user.username,
    )
@router.get("/export-full-data")
def export_full_data(
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
    allotment_date: str | None = Query(default=None),
    format: str = Query(default="json"),
    db: Session = Depends(get_db),
    _current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin, UserRole.user)),
):
    result = export_full_data_use_case(
        db,
        from_date=from_date,
        to_date=to_date,
        allotment_date=allotment_date,
        output_format=format,
    )

    if isinstance(result, ExportBinaryResult):
        return Response(
            content=result.content,
            media_type=result.media_type,
            headers={"Content-Disposition": f"attachment; filename={result.filename}"},
        )

    return result

    '''
    Legacy inline implementation (migrated to domain/repository; kept temporarily for reference).
    _ensure_claim_report_uploads_table(db)
    _ensure_claim_legacy_data_table(db)

    def _format_status(raw: str | None) -> str:
        v = str(raw or "").strip().lower()
        if v in {"waiting_for_documents", "ready_for_assignment", "pending"}:
            return "pending"
        return v or "pending"

    def _fmt_date(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%d-%m-%Y")
        if isinstance(v, date):
            return v.strftime("%d-%m-%Y")

        s = str(v).strip()
        if not s:
            return ""

        normalized = s.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(normalized)
            return dt.strftime("%d-%m-%Y")
        except Exception:
            pass

        date_formats = [
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d-%m-%Y %H:%M",
            "%d-%m-%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
        ]
        for fmt in date_formats:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%d-%m-%Y")
            except Exception:
                continue

        m = re.search(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", s)
        if m:
            token = m.group(0)
            for fmt in ("%d-%m-%Y", "%d/%m/%Y"):
                try:
                    return datetime.strptime(token, fmt).strftime("%d-%m-%Y")
                except Exception:
                    continue

        m = re.search(r"\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b", s)
        if m:
            token = m.group(0)
            for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
                try:
                    return datetime.strptime(token, fmt).strftime("%d-%m-%Y")
                except Exception:
                    continue

        return s

    def _fmt_datetime(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%d-%m-%Y %H:%M:%S")
        if isinstance(v, date):
            return datetime(v.year, v.month, v.day).strftime("%d-%m-%Y %H:%M:%S")

        s = str(v).strip()
        if not s:
            return ""

        normalized = s.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(normalized)
            return dt.strftime("%d-%m-%Y %H:%M:%S")
        except Exception:
            pass

        datetime_formats = [
            "%d-%m-%Y %H:%M:%S",
            "%d-%m-%Y %H:%M",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%Y/%m/%d",
        ]
        for fmt in datetime_formats:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%d-%m-%Y %H:%M:%S")
            except Exception:
                continue

        return s

    def _tag_at(tags: Any, idx: int) -> str:
        if isinstance(tags, list):
            if 0 <= idx < len(tags):
                return str(tags[idx] or "").strip()
            return ""
        return ""

    def _legacy_get(payload_obj: Any, *keys: str) -> str:
        if not isinstance(payload_obj, dict):
            return ""
        for key in keys:
            value = payload_obj.get(key)
            if value is None:
                continue
            text_value = str(value).strip()
            if text_value:
                return text_value
        return ""

    filters = []
    params: dict[str, Any] = {}

    if _is_valid_date(from_date):
        filters.append("DATE(c.created_at) >= :from_date")
        params["from_date"] = from_date
    if _is_valid_date(to_date):
        filters.append("b.allotment_date <= :to_date")
        params["to_date"] = to_date
    if _is_valid_date(allotment_date):
        filters.append("(EXISTS (SELECT 1 FROM workflow_events we WHERE we.claim_id = c.id AND we.event_type = 'claim_assigned' AND DATE(we.occurred_at) = :allotment_date) OR CASE WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{4}-\d{2}-\d{2}$' THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'YYYY-MM-DD') WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{2}-\d{2}-\d{4}$' THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD-MM-YYYY') WHEN NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), '') ~ '^\d{2}/\d{2}/\d{4}$' THEN TO_DATE(NULLIF(TRIM(COALESCE(ldata.legacy_payload->>'allocation_date', '')), ''), 'DD/MM/YYYY') ELSE NULL END = :allotment_date)")
        params["allotment_date"] = allotment_date

    where_sql = ("WHERE " + " AND ".join(filters)) if filters else ""

    rows = db.execute(
        text(
            f"""
            WITH latest_report AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id, report_status, export_uri, version_no, created_at AS report_created_at
                FROM report_versions
                ORDER BY claim_id, version_no DESC
            ),
            latest_assignment AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    DATE(occurred_at) AS allotment_date
                FROM workflow_events
                WHERE event_type = 'claim_assigned'
                ORDER BY claim_id, occurred_at DESC
            ),
            doc_stats AS (
                SELECT claim_id, COUNT(*) AS documents
                FROM claim_documents
                GROUP BY claim_id
            ),
            latest_decision AS (
                SELECT DISTINCT ON (claim_id)
                    claim_id,
                    explanation_summary
                FROM decision_results
                WHERE is_active = TRUE
                ORDER BY claim_id, generated_at DESC
            ),
            upload_meta AS (
                SELECT
                    claim_id,
                    report_export_status,
                    tagging,
                    subtagging,
                    opinion,
                    qc_status,
                    updated_by
                FROM claim_report_uploads
            ),
            legacy_data AS (
                SELECT claim_id, legacy_payload
                FROM claim_legacy_data
            )
            SELECT
                ROW_NUMBER() OVER (ORDER BY c.created_at DESC) AS row_id,
                c.external_claim_id,
                c.patient_name,
                c.patient_identifier,
                c.status,
                c.assigned_doctor_id,
                c.priority,
                c.source_channel,
                c.tags,
                c.created_at,
                c.updated_at,
                la.allotment_date,
                COALESCE(ds.documents, 0) AS documents,
                COALESCE(ldec.explanation_summary, '') AS trigger_remarks,
                COALESCE(rv.report_status, 'pending') AS report_status,
                COALESCE(rv.export_uri, '') AS export_uri,
                COALESCE(rv.version_no, 0) AS version_no,
                rv.report_created_at,
                COALESCE(um.report_export_status, CASE WHEN COALESCE(rv.export_uri, '') <> '' THEN 'uploaded' ELSE 'pending' END) AS report_export_status,
                COALESCE(um.tagging, '') AS tagging,
                COALESCE(um.subtagging, '') AS subtagging,
                COALESCE(um.opinion, '') AS opinion,
                CASE WHEN LOWER(REPLACE(REPLACE(COALESCE(um.qc_status, 'no'), ' ', '_'), '-', '_')) IN ('yes', 'qc_yes', 'qcyes', 'qc_done', 'done') THEN 'yes' ELSE 'no' END AS qc_status,
                COALESCE(um.updated_by, '') AS uploaded_by_username,
                u.id AS uploaded_by_user_id,
                ldata.legacy_payload
            FROM claims c
            LEFT JOIN latest_report rv ON rv.claim_id = c.id
            LEFT JOIN latest_assignment la ON la.claim_id = c.id
            LEFT JOIN doc_stats ds ON ds.claim_id = c.id
            LEFT JOIN latest_decision ldec ON ldec.claim_id = c.id
            LEFT JOIN upload_meta um ON um.claim_id = c.id
            LEFT JOIN users u ON u.username = um.updated_by
            LEFT JOIN legacy_data ldata ON ldata.claim_id = c.id
            {where_sql}
            ORDER BY c.created_at DESC
            """
        ),
        params,
    ).mappings().all()

    items = [
        {
            "external_claim_id": str(r.get("external_claim_id") or ""),
            "patient_name": str(r.get("patient_name") or ""),
            "patient_identifier": str(r.get("patient_identifier") or ""),
            "status": _format_status(r.get("status")),
            "assigned_doctor_id": str(r.get("assigned_doctor_id") or ""),
            "priority": int(r.get("priority") or 0),
            "source_channel": str(r.get("source_channel") or ""),
            "created_at": str(r.get("created_at") or ""),
            "updated_at": str(r.get("updated_at") or ""),
            "allotment_date": str(r.get("allotment_date") or ""),
            "report_status": str(r.get("report_status") or ""),
            "export_uri": str(r.get("export_uri") or ""),
            "report_version": int(r.get("version_no") or 0),
            "report_created_at": str(r.get("report_created_at") or ""),
        }
        for r in rows
    ]

    legacy_fieldnames = [
        "claim_date",
        "claim_id",
        "claim_type",
        "policy_number",
        "policy_type",
        "policy_start_date",
        "policy_end_date",
        "benef_name",
        "benef_age",
        "benef_gender",
        "pri_benef_name",
        "benef_sum_insured",
        "relation_type",
        "hospital_name",
        "hospital_pincode",
        "hospital_city",
        "hospital_state",
        "claim_amount",
        "doa_date",
        "dod_date",
        "claimant_ir",
        "hospital_is_network",
        "trigger_remarks",
        "document_required",
        "primary_icd_group",
        "primary_ailment_code",
        "treatment_type",
        "bill_deduction_reason",
        "vendor_name",
        "allocation_date",
        "document_status",
        "final_status",
        "report_export_status",
        "tagging",
        "subtagging",
        "opinion",
        "qc_status",
        "created_at",
        "updated_at",
    ]

    def _claim_id_number(value: Any) -> int | str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        digits = re.sub(r"[^0-9]", "", raw)
        if not digits:
            return raw
        try:
            return int(digits)
        except Exception:
            return raw

    legacy_items: list[dict[str, Any]] = []
    for r in rows:
        tags = r.get("tags")
        status_value = _format_status(r.get("status"))
        payload_obj = r.get("legacy_payload")
        if isinstance(payload_obj, str):
            try:
                payload_obj = json.loads(payload_obj)
            except Exception:
                payload_obj = {}
        if not isinstance(payload_obj, dict):
            payload_obj = {}


        legacy_items.append(
            {
                "claim_date": _fmt_date(_legacy_get(payload_obj, "claim_date") or r.get("created_at")),
                "claim_id": _claim_id_number(_legacy_get(payload_obj, "claim_id") or str(r.get("external_claim_id") or "")), 
                "claim_type": _legacy_get(payload_obj, "claim_type") or _tag_at(tags, 0),
                "policy_number": _legacy_get(payload_obj, "policy_number") or str(r.get("patient_identifier") or ""),
                "policy_type": _legacy_get(payload_obj, "policy_type") or _tag_at(tags, 1),
                "policy_start_date": _fmt_date(_legacy_get(payload_obj, "policy_start_date")),
                "policy_end_date": _fmt_date(_legacy_get(payload_obj, "policy_end_date")),
                "benef_name": _legacy_get(payload_obj, "benef_name") or str(r.get("patient_name") or ""),
                "benef_age": _legacy_get(payload_obj, "benef_age"),
                "benef_gender": _legacy_get(payload_obj, "benef_gender"),
                "pri_benef_name": _legacy_get(payload_obj, "pri_benef_name"),
                "benef_sum_insured": _legacy_get(payload_obj, "benef_sum_insured"),
                "relation_type": _legacy_get(payload_obj, "relation_type"),
                "hospital_name": _legacy_get(payload_obj, "hospital_name") or _tag_at(tags, 3),
                "hospital_pincode": _legacy_get(payload_obj, "hospital_pincode"),
                "hospital_city": _legacy_get(payload_obj, "hospital_city"),
                "hospital_state": _legacy_get(payload_obj, "hospital_state"),
                "claim_amount": _legacy_get(payload_obj, "claim_amount"),
                "doa_date": _fmt_date(_legacy_get(payload_obj, "doa_date", "doa", "doa date", "date_of_admission", "date of admission", "admission_date", "admission date")),
                "dod_date": _fmt_date(_legacy_get(payload_obj, "dod_date", "dod", "dod date", "date_of_discharge", "date of discharge", "discharge_date", "discharge date")),
                "claimant_ir": _legacy_get(payload_obj, "claimant_ir"),
                "hospital_is_network": _legacy_get(payload_obj, "hospital_is_network"),
                "trigger_remarks": _legacy_get(payload_obj, "trigger_remarks") or str(r.get("trigger_remarks") or ""),
                "document_required": _legacy_get(payload_obj, "document_required"),
                "primary_icd_group": _legacy_get(payload_obj, "primary_icd_group") or _tag_at(tags, 2),
                "primary_ailment_code": _legacy_get(payload_obj, "primary_ailment_code"),
                "treatment_type": _legacy_get(payload_obj, "treatment_type") or _tag_at(tags, 4),
                "bill_deduction_reason": _legacy_get(payload_obj, "bill_deduction_reason"),
                "vendor_name": _legacy_get(payload_obj, "vendor_name") or str(r.get("source_channel") or ""),
                "allocation_date": _fmt_date(_legacy_get(payload_obj, "allocation_date") or r.get("allotment_date")),
                "document_status": _legacy_get(payload_obj, "document_status") or ("uploaded" if int(r.get("documents") or 0) > 0 else "pending"),
                "final_status": _legacy_get(payload_obj, "final_status") or status_value,
                "report_export_status": _legacy_get(payload_obj, "report_export_status") or str(r.get("report_export_status") or "pending"),
                "tagging": _legacy_get(payload_obj, "tagging") or str(r.get("tagging") or ""),
                "subtagging": _legacy_get(payload_obj, "subtagging") or str(r.get("subtagging") or ""),
                "opinion": _legacy_get(payload_obj, "opinion") or str(r.get("opinion") or ""),
                "qc_status": _legacy_get(payload_obj, "qc_status") or str(r.get("qc_status") or "no"),
                "created_at": _fmt_datetime(_legacy_get(payload_obj, "created_at") or r.get("created_at")),
                "updated_at": _fmt_datetime(_legacy_get(payload_obj, "updated_at") or r.get("updated_at")),
            }
        )

    normalized_format = format.lower()
    if normalized_format == "csv":
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=legacy_fieldnames, extrasaction="ignore")
        writer.writeheader()
        for item in legacy_items:
            writer.writerow(item)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return PlainTextResponse(
            content=buf.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=user_full_data_{stamp}.csv"},
        )

    if normalized_format in {"excel", "xlsx"}:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            from openpyxl import Workbook
            from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

            def _xlsx_value(v: Any) -> Any:
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    return v
                raw = "" if v is None else str(v)
                cleaned = ILLEGAL_CHARACTERS_RE.sub("", raw)
                # XLSX cell text max is 32767 characters.
                if len(cleaned) > 32767:
                    cleaned = cleaned[:32767]
                return cleaned

            wb = Workbook()
            ws = wb.active
            ws.title = "user_full_data"
            ws.append(legacy_fieldnames)
            for item in legacy_items:
                ws.append([_xlsx_value(item.get(col, "")) for col in legacy_fieldnames])

            out = io.BytesIO()
            wb.save(out)
            return Response(
                content=out.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=user_full_data_{stamp}.xlsx"},
            )
        except Exception:
            buf = io.StringIO()
            writer = csv.DictWriter(buf, fieldnames=legacy_fieldnames, extrasaction="ignore")
            writer.writeheader()
            for item in legacy_items:
                writer.writerow(item)

            return PlainTextResponse(
                content=buf.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=user_full_data_{stamp}.csv"},
            )

    '''

@router.get("/payment-sheet")
def payment_sheet(
    month: str | None = Query(default=None),
    include_zero_cases: bool = Query(default=True),
    db: Session = Depends(get_db),
    _current_user: AuthenticatedUser = Depends(require_roles(UserRole.super_admin)),
) -> dict:
    try:
        return get_payment_sheet_use_case(
            db,
            month=month,
            include_zero_cases=bool(include_zero_cases),
        )
    except PaymentSheetInvalidMonthError as exc:
        raise HTTPException(status_code=400, detail=exc.message) from exc
