from __future__ import annotations

import csv
import io
import re
from typing import Any

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.repositories import claim_legacy_data_repo, claim_report_uploads_repo, claims_repo
from app.services.sql_dump_parser import iter_table_rows_from_sql_dump_bytes


_EMPTY_LIKE_TEXT_VALUES = {
    "na",
    "n/a",
    "not available",
    "none",
    "nil",
    "null",
    "-",
    ".",
}


def _normalize_import_key(value: Any) -> str:
    key = str(value or "").strip().lower()
    if not key:
        return ""
    key = key.replace("\ufeff", "")
    key = re.sub(r"[\s\-/]+", "_", key)
    key = re.sub(r"[^a-z0-9_]", "", key)
    key = re.sub(r"_+", "_", key).strip("_")
    return key


def _row_with_normalized_aliases(raw_map: dict[Any, Any]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in (raw_map or {}).items():
        raw_key = str(key or "").strip().lower()
        if not raw_key:
            continue
        text_value = str(value or "").strip()
        normalized[raw_key] = text_value
        normalized_key = _normalize_import_key(raw_key)
        if normalized_key and normalized_key not in normalized:
            normalized[normalized_key] = text_value
    return normalized


def _parse_csv_rows(payload: bytes) -> list[dict[str, str]]:
    text_payload = payload.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text_payload))
    return [_row_with_normalized_aliases(row) for row in reader]


def _parse_xlsx_rows(payload: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed for xlsx import") from exc

    wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(h or "").strip().lower() for h in headers_raw]
    rows: list[dict[str, str]] = []
    for row_vals in rows_iter:
        raw_row: dict[str, str] = {}
        for idx, val in enumerate(row_vals):
            key = headers[idx] if idx < len(headers) else f"col_{idx + 1}"
            raw_row[key] = str(val).strip() if val is not None else ""
        rows.append(_row_with_normalized_aliases(raw_row))
    return rows


def _parse_sql_dump_rows(payload: bytes) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for raw_row in iter_table_rows_from_sql_dump_bytes(payload, "excel_case_uploads"):
        prepared_row: dict[str, str] = {}
        for key, value in (raw_row or {}).items():
            raw_key = str(key or "").strip().lower()
            if not raw_key:
                continue
            if value is None:
                prepared_row[raw_key] = ""
            elif isinstance(value, bool):
                prepared_row[raw_key] = "1" if value else "0"
            else:
                prepared_row[raw_key] = str(value).strip()
        normalized = _row_with_normalized_aliases(prepared_row)
        if normalized:
            rows.append(normalized)
    return rows


def parse_import_rows(*, filename: str, payload: bytes) -> list[dict[str, str]]:
    name = str(filename or "").lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx_rows(payload)
    if name.endswith(".sql"):
        return _parse_sql_dump_rows(payload)
    return _parse_csv_rows(payload)


def _coerce_claim_status(raw: str | None) -> str:
    v = (raw or "").strip().lower()
    valid = {
        "ready_for_assignment",
        "waiting_for_documents",
        "in_review",
        "needs_qc",
        "completed",
        "withdrawn",
    }
    return v if v in valid else "waiting_for_documents"


def extract_claim_fields(row: dict[str, str]) -> dict[str, Any] | None:
    external_claim_id = (
        row.get("external_claim_id")
        or row.get("claim_id")
        or row.get("claim no")
        or row.get("claim")
        or ""
    ).strip()
    if not external_claim_id:
        return None

    patient_name = (row.get("patient_name") or row.get("benef_name") or row.get("patient") or "").strip() or None

    patient_identifier = (
        row.get("patient_identifier") or row.get("policy_number") or row.get("policy no") or ""
    ).strip() or None

    assigned_doctor_id = (
        row.get("assigned_doctor_id") or row.get("doctor_username") or row.get("doctor") or ""
    ).strip() or None

    status = _coerce_claim_status(row.get("status"))
    if assigned_doctor_id and status == "waiting_for_documents":
        status = "in_review"

    tags: list[str] = []
    for key in ["claim_type", "policy_type", "diagnosis", "hospital_name", "treatment_type"]:
        val = (row.get(key) or "").strip()
        if val:
            tags.append(val)

    return {
        "external_claim_id": external_claim_id,
        "patient_name": patient_name,
        "patient_identifier": patient_identifier,
        "status": status,
        "assigned_doctor_id": assigned_doctor_id,
        "priority": 3,
        "source_channel": "excel_import",
        "tags": tags,
    }


def import_claims_from_excel_payload(
    db: Session,
    *,
    payload: bytes,
    filename: str,
    uploaded_by_username: str,
) -> tuple[int, int, int, int]:
    claim_legacy_data_repo.ensure_table(db)
    claim_report_uploads_repo.ensure_claim_report_uploads_table(db)
    claims_repo.ensure_claim_completed_at_column_and_backfill(db)

    rows = parse_import_rows(filename=filename, payload=payload)

    inserted = 0
    updated = 0
    skipped = 0

    for row in rows:
        claim = extract_claim_fields(row)
        if claim is None:
            skipped += 1
            continue

        existing = claims_repo.get_claim_row_by_external_claim_id(db, external_claim_id=str(claim["external_claim_id"]))

        if existing is None:
            created = claims_repo.insert_claim_from_integration(
                db,
                external_claim_id=str(claim["external_claim_id"]),
                patient_name=claim.get("patient_name"),
                patient_identifier=claim.get("patient_identifier"),
                status=str(claim["status"]),
                assigned_doctor_id=claim.get("assigned_doctor_id"),
                priority=int(claim.get("priority") or 3),
                source_channel=str(claim.get("source_channel") or "excel_import"),
                tags=list(claim.get("tags") or []),
            )
            claim_uuid = str(created["id"])
            inserted += 1
        else:
            claim_uuid = str(existing["id"])
            claims_repo.update_claim_from_integration(
                db,
                claim_id=claim_uuid,
                patient_name=str(claim.get("patient_name") or ""),
                patient_identifier=str(claim.get("patient_identifier") or ""),
                assigned_doctor_id=str(claim.get("assigned_doctor_id") or ""),
                status=str(claim["status"]),
                priority=int(claim.get("priority") or 3),
                source_channel=str(claim.get("source_channel") or "excel_import"),
                tags=list(claim.get("tags") or []),
            )
            updated += 1

        legacy_payload = dict(row)
        if not str(legacy_payload.get("claim_id") or "").strip():
            legacy_payload["claim_id"] = str(claim.get("external_claim_id") or "")
        if not str(legacy_payload.get("source_file_name") or "").strip():
            legacy_payload["source_file_name"] = str(filename or "")
        if not str(legacy_payload.get("uploaded_by_username") or "").strip():
            legacy_payload["uploaded_by_username"] = str(uploaded_by_username or "")

        claim_legacy_data_repo.upsert_legacy_data(db, claim_id=claim_uuid, legacy_payload=legacy_payload)

    db.commit()
    return (len(rows), inserted, updated, skipped)


__all__ = ["import_claims_from_excel_payload"]

